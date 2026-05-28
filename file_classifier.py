"""
文件自动归类程序
================
功能：
  1. 实时递归监视指定文件夹，文件写入完成后将其复制到输出文件夹
  2. 按程序同目录下的发票号清单 Excel 第一列发票号匹配文件名前缀，自动归入对应发票号子文件夹
  3. 识别退税材料齐套状态，生成判断报告，并按报关号改名发票号文件夹
  4. 输出主文件夹生成以日期命名的 JSON / Excel 工作日志
  5. 图形界面，适合普通用户操作
  6. 支持批量文件处理（后台队列处理，避免界面卡死）

依赖安装：
  pip install watchdog openpyxl PyMuPDF rapidocr-onnxruntime
"""

from __future__ import annotations

import os
import re
import sys
import time
import shutil
import logging
import threading
import queue
import json
import tempfile
import unicodedata
import argparse
import filecmp
import hashlib
from datetime import date, datetime, timedelta
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext, simpledialog

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    Observer = None
    FileSystemEventHandler = object

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    openpyxl = None


# ─────────────────────────────────────────────
# 工具函数
# ─────────────────────────────────────────────

SCRIPT_DIR = (
    Path(sys.executable).resolve().parent
    if getattr(sys, "frozen", False)
    else Path(__file__).resolve().parent
)
DEFAULT_INVOICE_WORKBOOK = SCRIPT_DIR / "04月发票号清单.xlsx"
APP_SETTINGS_PATH = SCRIPT_DIR / "app_settings.json"
DATA_DIR = SCRIPT_DIR / "data"
DEFAULT_COMPANY_LEDGER = DATA_DIR / "公司系统出运统计发票号清单.xlsx"
DEFAULT_MATERIAL_CONFIG = SCRIPT_DIR / "materials_config.json"
REVIEW_SUBFOLDER = "未匹配发票号"
IGNORED_SUFFIXES = {
    ".tmp", ".temp", ".crdownload", ".part", ".download", ".swp", ".lock",
}
IGNORED_PREFIXES = ("~$", ".~lock")
REPORT_PREFIXES = ("工作日志_", "判断报告_", "rename_manifest_", "operation_manifest_")
IMAGE_OCR_SUFFIXES = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}
ROUTABLE_MATERIAL_STATUSES = {"matched", "manual_review"}
STATUS_SKIPPED = "跳过"
STATUS_PLANNED = "预览"
DEFAULT_DECLARATION_TIMEOUT_SECONDS = 60.0
SAFE_BATCH_CONFIRM_LIMIT = 20
MAX_FILE_PROCESS_WORKERS = 4
OPERATION_MANIFEST_SCHEMA = "contract-router-operation-manifest/v1"
OPERATION_MANIFEST_DIRNAME = "_操作日志"
OPERATION_MANIFEST_PREFIX = "operation_manifest_"
EXCEL_WRITE_LOCK = threading.Lock()


def default_app_settings() -> dict:
    return {
        "company_ledger": {
            "path": "data/公司系统出运统计发票号清单.xlsx"
        },
        "finance_batch": {
            "output_subdir": "财务退税批次包",
            "copy_mode": "copy",
            "write_back_company_ledger": True,
            "write_back_finance_batch": True,
            "create_marked_batch_copy": True
        }
    }


def _merge_missing_settings(base: dict, defaults: dict) -> bool:
    changed = False
    for key, value in defaults.items():
        if key not in base:
            base[key] = value
            changed = True
            continue
        if isinstance(base.get(key), dict) and isinstance(value, dict):
            changed = _merge_missing_settings(base[key], value) or changed
    return changed


def ensure_app_settings(path: Path = APP_SETTINGS_PATH) -> dict:
    defaults = default_app_settings()
    path = Path(path)
    if not path.exists():
        path.parent.mkdir(parents=True, exist_ok=True)
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(defaults, f, ensure_ascii=False, indent=2)
        return defaults

    try:
        with open(path, "r", encoding="utf-8") as f:
            settings = json.load(f)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"app_settings.json 格式错误，请修正后重试：{path}") from exc

    if not isinstance(settings, dict):
        settings = {}
    if _merge_missing_settings(settings, defaults):
        save_app_settings(settings, path)
    return settings


def save_app_settings(settings: dict, path: Path = APP_SETTINGS_PATH):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(settings, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def resolve_app_path(relative_or_abs: str) -> Path:
    raw = str(relative_or_abs or "").strip()
    if not raw:
        return DEFAULT_COMPANY_LEDGER
    path = Path(raw).expanduser()
    if path.is_absolute() or re.match(r"^[A-Za-z]:[\\/]", raw):
        return path
    return SCRIPT_DIR / path


def normalize_path(path: str) -> str:
    return os.path.abspath(os.path.normpath(path))


def comparable_path(path: str) -> str:
    return os.path.normcase(normalize_path(path))


def paths_equal(left: str, right: str) -> bool:
    return comparable_path(left) == comparable_path(right)


def is_path_inside(child: str, parent: str) -> bool:
    child_cmp = comparable_path(child)
    parent_cmp = comparable_path(parent)
    try:
        return os.path.commonpath([child_cmp, parent_cmp]) == parent_cmp
    except ValueError:
        return False


def should_ignore_file(filepath: str) -> bool:
    filename = os.path.basename(filepath)
    suffix = Path(filename).suffix.lower()
    return (
        not filename
        or filename.startswith(IGNORED_PREFIXES)
        or filename.startswith(REPORT_PREFIXES)
        or suffix in IGNORED_SUFFIXES
    )


def sanitize_folder_name(name: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|]', "_", str(name).strip()).strip(". ")
    return cleaned or "未分类"


def unique_path(path: str) -> str:
    if not os.path.exists(path):
        return path

    folder = os.path.dirname(path)
    stem = Path(path).stem
    suffix = Path(path).suffix
    counter = 1
    while True:
        candidate = os.path.join(folder, f"{stem}_重复{counter}{suffix}")
        if not os.path.exists(candidate):
            return candidate
        counter += 1


def files_have_same_content(first_path: str, second_path: str) -> bool:
    try:
        if not os.path.isfile(first_path) or not os.path.isfile(second_path):
            return False
        if os.path.getsize(first_path) != os.path.getsize(second_path):
            return False
        return filecmp.cmp(first_path, second_path, shallow=False)
    except OSError:
        return False


def file_sha256(path: str) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def file_state(path: str) -> dict:
    if not path or not os.path.isfile(path):
        return {}
    try:
        return {
            "size": os.path.getsize(path),
            "sha256": file_sha256(path),
        }
    except OSError:
        return {}


def operation_manifest_dir(output_dir: str) -> str:
    return os.path.join(output_dir, OPERATION_MANIFEST_DIRNAME)


def operation_manifest_path(output_dir: str, run_id: str = "") -> str:
    run_id = sanitize_folder_name(run_id) if run_id else datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(
        operation_manifest_dir(output_dir),
        f"{OPERATION_MANIFEST_PREFIX}{run_id}_{os.getpid()}.json",
    )


def list_operation_manifests(output_dir: str) -> list[str]:
    folder = operation_manifest_dir(output_dir)
    if not os.path.isdir(folder):
        return []
    paths = []
    for name in os.listdir(folder):
        if name.startswith(OPERATION_MANIFEST_PREFIX) and name.endswith(".json"):
            paths.append(os.path.join(folder, name))
    return sorted(paths, key=lambda item: os.path.getmtime(item), reverse=True)


def latest_operation_manifest(output_dir: str) -> str:
    manifests = list_operation_manifests(output_dir)
    return manifests[0] if manifests else ""


class OperationManifest:
    def __init__(self, output_dir: str, dry_run: bool = False, run_id: str = "", path: str = ""):
        self.output_dir = normalize_path(output_dir)
        self.dry_run = bool(dry_run)
        self.path = normalize_path(path or operation_manifest_path(self.output_dir, run_id))
        self.lock = threading.Lock()
        self.data = {
            "schema": OPERATION_MANIFEST_SCHEMA,
            "run_id": run_id or Path(self.path).stem.removeprefix(OPERATION_MANIFEST_PREFIX),
            "created_at": datetime.now().isoformat(timespec="seconds"),
            "output_dir": self.output_dir,
            "dry_run": self.dry_run,
            "operations": [],
        }
        os.makedirs(os.path.dirname(self.path), exist_ok=True)
        self._save_locked()

    def add(
        self,
        action: str,
        source_path: str = "",
        destination_path: str = "",
        status: str = "",
        invoice_no: str = "",
        subfolder: str = "",
        message: str = "",
        extra: dict | None = None,
    ) -> dict:
        record = {
            "id": f"{time.time_ns()}-{threading.get_ident()}",
            "time": datetime.now().isoformat(timespec="seconds"),
            "action": action,
            "status": status,
            "source_path": normalize_path(source_path) if source_path else "",
            "destination_path": normalize_path(destination_path) if destination_path else "",
            "invoice_no": invoice_no,
            "subfolder": subfolder,
            "message": message,
        }
        if extra:
            record.update(extra)
        with self.lock:
            self.data["operations"].append(record)
            self._save_locked()
        return record

    def _save_locked(self) -> None:
        tmp_path = f"{self.path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, self.path)


def resolve_copy_destination(src_path: str, dest_path: str) -> tuple[str, bool]:
    if not os.path.exists(dest_path):
        return dest_path, False
    if files_have_same_content(src_path, dest_path):
        return dest_path, True

    folder = os.path.dirname(dest_path)
    stem = Path(dest_path).stem
    suffix = Path(dest_path).suffix
    counter = 1
    while True:
        candidate = os.path.join(folder, f"{stem}_重复{counter}{suffix}")
        if os.path.exists(candidate):
            if files_have_same_content(src_path, candidate):
                return candidate, True
            counter += 1
            continue
        return candidate, False


def record_file_copy(
    operation_manifest: OperationManifest | None,
    action: str,
    source_path: str,
    destination_path: str,
    status: str,
    invoice_no: str = "",
    subfolder: str = "",
    message: str = "",
    planned: bool = False,
) -> None:
    if not operation_manifest:
        return
    extra = {"planned": bool(planned)}
    source_state = file_state(source_path)
    destination_state = file_state(destination_path)
    if source_state:
        extra["source_state"] = source_state
    if destination_state:
        extra["destination_state"] = destination_state
    operation_manifest.add(
        action=action,
        source_path=source_path,
        destination_path=destination_path,
        status=status,
        invoice_no=invoice_no,
        subfolder=subfolder,
        message=message,
        extra=extra,
    )


def copy_file_with_manifest(
    src_path: str,
    dest_path: str,
    operation_manifest: OperationManifest | None = None,
    action: str = "copy_file",
    invoice_no: str = "",
    subfolder: str = "",
    message: str = "",
    dry_run: bool = False,
) -> tuple[str, bool]:
    dest_path, reused = resolve_copy_destination(src_path, dest_path)
    status = "planned_reuse" if dry_run and reused else "planned_copy" if dry_run else "reused" if reused else "copied"
    if dry_run:
        record_file_copy(
            operation_manifest,
            action,
            src_path,
            dest_path,
            status,
            invoice_no=invoice_no,
            subfolder=subfolder,
            message=message,
            planned=True,
        )
        return dest_path, reused
    if not reused:
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        shutil.copy2(src_path, dest_path)
    record_file_copy(
        operation_manifest,
        action,
        src_path,
        dest_path,
        status,
        invoice_no=invoice_no,
        subfolder=subfolder,
        message=message,
    )
    return dest_path, reused


def copy_folder_contents(
    source_dir: str,
    target_dir: str,
    operation_manifest: OperationManifest | None = None,
    action: str = "copy_folder_file",
    dry_run: bool = False,
) -> int:
    source_dir = normalize_path(source_dir)
    target_dir = normalize_path(target_dir)
    if paths_equal(source_dir, target_dir) or not os.path.isdir(source_dir):
        return 0
    copied = 0
    for root, _, files in os.walk(source_dir):
        rel_root = os.path.relpath(root, source_dir)
        dest_root = target_dir if rel_root == "." else os.path.join(target_dir, rel_root)
        for filename in files:
            src = os.path.join(root, filename)
            dst = os.path.join(dest_root, filename)
            dst, reused = copy_file_with_manifest(
                src,
                dst,
                operation_manifest=operation_manifest,
                action=action,
                message=f"复制合并: {src} -> {dst}",
                dry_run=dry_run,
            )
            if not reused:
                copied += 1
    return copied


def merge_folder_contents(source_dir: str, target_dir: str) -> None:
    copy_folder_contents(source_dir, target_dir)


def remove_empty_parents(path: str, stop_dir: str) -> None:
    stop_dir = comparable_path(stop_dir)
    current = os.path.dirname(normalize_path(path))
    while current and comparable_path(current) != stop_dir:
        try:
            os.rmdir(current)
        except OSError:
            return
        current = os.path.dirname(current)


def undo_operation_manifest(manifest_path: str, dry_run: bool = False) -> dict:
    manifest_path = normalize_path(manifest_path)
    with open(manifest_path, "r", encoding="utf-8") as f:
        manifest = json.load(f)
    if manifest.get("schema") != OPERATION_MANIFEST_SCHEMA:
        raise RuntimeError(f"不支持的操作清单格式：{manifest_path}")

    output_dir = normalize_path(manifest.get("output_dir") or os.path.dirname(manifest_path))
    summary = {
        "manifest_path": manifest_path,
        "dry_run": bool(dry_run),
        "total": 0,
        "undone": 0,
        "skipped": 0,
        "errors": [],
        "details": [],
    }
    operations = manifest.get("operations", [])
    if not isinstance(operations, list):
        raise RuntimeError(f"操作清单损坏：{manifest_path}")

    for operation in reversed(operations):
        if not isinstance(operation, dict):
            continue
        summary["total"] += 1
        action = operation.get("action", "")
        status = operation.get("status", "")
        src = operation.get("source_path", "")
        dst = operation.get("destination_path", "")

        try:
            if status == "copied" and dst and action in {"copy_file", "copy_merge_file", "rename_copy_file"}:
                if not os.path.exists(dst):
                    summary["skipped"] += 1
                    summary["details"].append(f"已不存在，跳过: {dst}")
                    continue
                if not os.path.isfile(dst):
                    summary["skipped"] += 1
                    summary["details"].append(f"不是文件复制记录，跳过: {dst}")
                    continue
                expected_state = operation.get("destination_state") or {}
                current_state = file_state(dst)
                if (
                    expected_state.get("sha256")
                    and current_state.get("sha256")
                    and expected_state["sha256"] != current_state["sha256"]
                ):
                    summary["skipped"] += 1
                    summary["details"].append(f"目标已被改动，跳过: {dst}")
                    continue
                if not dry_run:
                    os.remove(dst)
                    remove_empty_parents(dst, output_dir)
                summary["undone"] += 1
                summary["details"].append(f"删除已复制文件: {dst}")
                continue

            if action == "rename_file" and status == "renamed" and src and dst:
                if not os.path.exists(dst):
                    summary["skipped"] += 1
                    summary["details"].append(f"改名后文件不存在，跳过: {dst}")
                    continue
                if os.path.exists(src):
                    summary["skipped"] += 1
                    summary["details"].append(f"原路径已有文件，跳过回退: {src}")
                    continue
                if not dry_run:
                    os.makedirs(os.path.dirname(src), exist_ok=True)
                    os.rename(dst, src)
                summary["undone"] += 1
                summary["details"].append(f"回退文件改名: {dst} -> {src}")
                continue

            if action == "rename_folder" and status == "renamed" and src and dst:
                if not os.path.isdir(dst):
                    summary["skipped"] += 1
                    summary["details"].append(f"改名后文件夹不存在，跳过: {dst}")
                    continue
                if os.path.exists(src):
                    summary["skipped"] += 1
                    summary["details"].append(f"原文件夹仍存在，跳过回退: {src}")
                    continue
                if not dry_run:
                    os.rename(dst, src)
                summary["undone"] += 1
                summary["details"].append(f"回退文件夹改名: {dst} -> {src}")
                continue

            summary["skipped"] += 1
        except Exception as exc:
            summary["errors"].append(f"{action or 'operation'}: {exc}")

    return summary


def undo_latest_operation_manifest(output_dir: str, dry_run: bool = False) -> dict:
    manifest_path = latest_operation_manifest(output_dir)
    if not manifest_path:
        raise FileNotFoundError(f"没有找到可撤销的操作清单：{operation_manifest_dir(output_dir)}")
    return undo_operation_manifest(manifest_path, dry_run=dry_run)


def normalize_unicode_text(value: str) -> str:
    return unicodedata.normalize("NFKC", str(value or ""))


def join_wrapped_identifiers(text: str) -> str:
    text = normalize_unicode_text(text)
    pattern = re.compile(
        r"\b([A-Za-z0-9]{6,})[ \t]*[\r\n]+[ \t]*"
        r"([A-Za-z0-9]{1,4})(?![ \t]*(?:%|CARTONS|CTNS|PACKAGES|KGS|PCS|PALLETS)\b)(?=\b)"
    )
    label_fragments = {
        "SEAL", "PORT", "DATE", "TOTAL", "CARGO", "MEAS", "PACK", "NO", "NOS",
        "JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "SEPT",
        "OCT", "NOV", "DEC",
    }

    def join_or_space(match: re.Match) -> str:
        left, right = match.group(1), match.group(2)
        left_upper = left.upper()
        right_upper = right.upper()
        if right_upper in label_fragments:
            return f"{left} {right}"
        if left_upper.endswith(("CARTONS", "CTNS", "PACKAGES", "KGS", "CBM", "PCS", "PALLETS")):
            return f"{left} {right}"
        if re.fullmatch(r"\d{2}H[QC]", right_upper):
            return f"{left} {right}"
        if re.search(r"\d", left) and (right.isdigit() or right.isupper()):
            return f"{left}{right}"
        return f"{left} {right}"

    previous = None
    while previous != text:
        previous = text
        text = pattern.sub(join_or_space, text)
    return text


def normalize_document_text(value: str) -> str:
    text = join_wrapped_identifiers(value)
    text = text.replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_text_for_match(value: str) -> str:
    return normalize_document_text(value).casefold()


def strip_invoice_prefix(filename: str, invoice_no: str) -> str:
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    prefixes = [invoice_no]
    digits = invoice_digits(invoice_no)
    if digits and digits != invoice_no:
        prefixes.append(digits)
    for prefix in prefixes:
        if prefix and re.match(re.escape(prefix) + r"(?![A-Za-z0-9])", stem, re.IGNORECASE):
            stem = stem[len(prefix):]
            stem = re.sub(r"^[\s_\-—–+]+", "", stem)
            break
    return f"{stem}{suffix}" if stem else filename


def invoice_display_value(invoice_no: str, mode: str = "digits_only") -> str:
    if mode == "digits_only":
        digits = "".join(re.findall(r"\d+", invoice_no))
        return digits or invoice_no
    return invoice_no


def format_excel_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_identifier(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", str(value or "")).upper()


def invoice_digits(value: str) -> str:
    return "".join(re.findall(r"\d+", str(value or "")))


def invoice_equivalent(left: str, right: str) -> bool:
    left_norm = normalize_identifier(left)
    right_norm = normalize_identifier(right)
    if not left_norm or not right_norm:
        return False
    if left_norm == right_norm:
        return True
    left_digits = invoice_digits(left_norm)
    right_digits = invoice_digits(right_norm)
    return bool(left_digits and len(left_digits) >= 6 and left_digits == right_digits)


def invoice_dedup_key(invoice_no: str) -> str:
    digits = invoice_digits(invoice_no)
    if len(digits) >= 6:
        return f"digits:{digits}"
    normalized = normalize_identifier(invoice_no)
    return f"id:{normalized}" if normalized else str(invoice_no or "").strip().casefold()


def normalize_amount(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value)
    text = re.sub(r"[^\d.\-]", "", text.replace(",", ""))
    if not text:
        return None
    try:
        return round(float(text), 2)
    except ValueError:
        return None


def amounts_equal(left, right, tolerance: float = 0.05) -> bool:
    left_amount = normalize_amount(left)
    right_amount = normalize_amount(right)
    if left_amount is None or right_amount is None:
        return False
    return abs(left_amount - right_amount) <= tolerance


MONTH_MAP = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


COUNTRY_MAP = {
    "POLAND": "波兰",
    "GDANSK": "波兰",
    "GERMANY": "德国",
    "ITALY": "意大利",
    "ECUADOR": "厄瓜多尔",
    "BELGIUM": "比利时",
    "LUXEMBOURG": "卢森堡",
    "UNITED STATES": "美国",
    "USA": "美国",
    "U.S.A": "美国",
}


def normalize_country(value: str) -> str:
    text = normalize_text_for_match(value)
    if not text:
        return ""
    for key, country in COUNTRY_MAP.items():
        if key.casefold() in text:
            return country
    return str(value).strip()


def normalize_date_value(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return (datetime(1899, 12, 30) + timedelta(days=float(value))).date().isoformat()
        except Exception:
            return ""

    text = str(value).strip()
    if not text:
        return ""

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y年%m月%d日", "%b.%d,%Y", "%b %d %Y", "%B %d %Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            pass

    match = re.search(
        r"\b(JANUARY|JAN|FEBRUARY|FEB|MARCH|MAR|APRIL|APR|MAY|JUNE|JUN|JULY|JUL|AUGUST|AUG|SEPTEMBER|SEPT|SEP|OCTOBER|OCT|NOVEMBER|NOV|DECEMBER|DEC)\.?\s*,?\s*(\d{1,2})\s*,?\s*(\d{4})\b",
        text,
        re.IGNORECASE,
    )
    if match:
        month = MONTH_MAP[match.group(1).upper()]
        day = int(match.group(2))
        year = int(match.group(3))
        return date(year, month, day).isoformat()
    return ""


def dates_close(left: str, right: str, max_days: int = 3) -> bool:
    try:
        left_date = datetime.fromisoformat(left).date()
        right_date = datetime.fromisoformat(right).date()
    except Exception:
        return False
    return abs((left_date - right_date).days) <= max_days


def find_invoice_workbook(settings: dict | None = None):
    settings = settings or ensure_app_settings(APP_SETTINGS_PATH)
    ledger_path = resolve_app_path(
        settings.get("company_ledger", {}).get(
            "path",
            "data/公司系统出运统计发票号清单.xlsx",
        )
    )
    if ledger_path.exists():
        return ledger_path
    if DEFAULT_INVOICE_WORKBOOK.exists():
        return DEFAULT_INVOICE_WORKBOOK
    return None


def load_invoice_numbers(workbook_path: str) -> list[str]:
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法读取发票号清单")

    invoices = []
    seen = set()
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        ws = wb.active
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            invoice_no = format_excel_cell(row[0])
            if not invoice_no or invoice_no in {"发票号", "invoice", "Invoice"}:
                continue
            key = invoice_no.casefold()
            if key in seen:
                continue
            seen.add(key)
            invoices.append(invoice_no)
    finally:
        wb.close()

    invoices.sort(key=len, reverse=True)
    return invoices


def load_invoice_records(workbook_path: str) -> dict[str, dict]:
    """读取发票号台账，并保留闭环校验需要的辅助字段。"""
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法读取发票号清单")

    records: dict[str, dict] = {}
    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = ws.iter_rows(values_only=True)
            try:
                headers = [format_excel_cell(value) for value in next(rows)]
            except StopIteration:
                continue

            header_index = {header: idx for idx, header in enumerate(headers) if header}
            invoice_idx = header_index.get("发票号")
            if invoice_idx is None:
                continue

            date_idx = header_index.get("出运日期")
            country_idx = (
                header_index.get("目的国(地区)")
                if "目的国(地区)" in header_index
                else header_index.get("目的国")
            )
            amount_idx = header_index.get("金额")

            for row_no, row in enumerate(rows, start=2):
                invoice_no = format_excel_cell(row[invoice_idx] if invoice_idx < len(row) else "")
                if not invoice_no:
                    continue
                records[invoice_no] = {
                    "发票号": invoice_no,
                    "sheet": ws.title,
                    "row": row_no,
                    "出运日期": normalize_date_value(row[date_idx] if date_idx is not None and date_idx < len(row) else None),
                    "目的国": normalize_country(row[country_idx] if country_idx is not None and country_idx < len(row) else ""),
                    "金额": normalize_amount(row[amount_idx] if amount_idx is not None and amount_idx < len(row) else None),
                }
    finally:
        wb.close()

    return records


class InvoiceMatcher:
    def __init__(self, invoice_numbers: list[str]):
        normalized = [
            str(invoice_no).strip()
            for invoice_no in invoice_numbers
            if str(invoice_no).strip()
        ]
        digit_counts = {}
        for invoice_no in normalized:
            digits = invoice_digits(invoice_no)
            if digits:
                digit_counts[digits] = digit_counts.get(digits, 0) + 1

        entries = []
        seen = set()
        for invoice_no in sorted(normalized, key=len, reverse=True):
            canonical = invoice_no.upper()
            for alias in (invoice_no, invoice_digits(invoice_no)):
                if not alias or (alias != invoice_no and digit_counts.get(alias) != 1):
                    continue
                key = alias.casefold()
                if key in seen:
                    continue
                seen.add(key)
                entries.append((key, canonical))
        self._entries = entries
        self._invoice_numbers = [invoice_no.upper() for invoice_no in normalized]

    def match(self, filename: str) -> str:
        raw_name = Path(filename).name.strip()
        normalized_name = normalize_unicode_text(raw_name)
        if "采购合同" in normalized_name or "合同" in normalized_name:
            return self._match_purchase_contract(normalized_name)

        name = normalized_name.casefold()
        for key, invoice_no in self._entries:
            if re.match(re.escape(key) + r"(?![A-Za-z0-9])", name, re.IGNORECASE):
                return invoice_no
        return ""

    def _match_purchase_contract(self, filename: str) -> str:
        digit_groups = re.findall(r"\d{6,12}", filename)
        if not digit_groups:
            return ""

        matches = {
            invoice_no
            for invoice_no in self._invoice_numbers
            for digits in digit_groups
            if invoice_equivalent(digits, invoice_no)
        }
        if len(matches) == 1:
            return next(iter(matches))
        return ""

    def match_path(self, filepath: str, source_root: str = "") -> str:
        invoice_no = self.match(Path(filepath).name)
        if invoice_no:
            return invoice_no

        path = Path(normalize_path(filepath))
        if source_root:
            try:
                relative_path = path.relative_to(Path(normalize_path(source_root)))
            except ValueError:
                relative_path = path
        else:
            relative_path = path

        for folder_name in reversed(relative_path.parts[:-1]):
            invoice_no = self.match(folder_name)
            if invoice_no:
                return invoice_no
        return ""


def can_acquire_file_lock(filepath: str) -> bool:
    """尝试取得独占文件锁，失败说明文件可能仍在写入或被其他程序占用。"""
    try:
        with open(filepath, "r+b") as fh:
            fh.seek(0)
            if os.name == "nt":
                import msvcrt
                try:
                    msvcrt.locking(fh.fileno(), msvcrt.LK_NBLCK, 1)
                    msvcrt.locking(fh.fileno(), msvcrt.LK_UNLCK, 1)
                    return True
                except OSError:
                    return False
            else:
                import fcntl
                try:
                    fcntl.flock(fh.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                    fcntl.flock(fh.fileno(), fcntl.LOCK_UN)
                    return True
                except OSError:
                    return False
    except OSError:
        return False


def wait_for_file_ready(
    filepath: str,
    timeout_seconds: float = 30.0,
    poll_interval: float = 1.0,
    stable_checks: int = 2,
    stop_event: threading.Event = None,
) -> tuple[bool, str]:
    """等待文件大小连续稳定，并且能取得独占锁。"""
    deadline = time.monotonic() + timeout_seconds
    last_size = None
    stable_count = 0
    last_reason = "文件尚未稳定"

    while time.monotonic() <= deadline:
        if stop_event and stop_event.is_set():
            return False, "监控已停止"
        if not os.path.exists(filepath):
            return False, "文件不存在"
        if should_ignore_file(filepath):
            return False, "临时文件或锁文件已忽略"

        try:
            current_size = os.path.getsize(filepath)
        except OSError as exc:
            last_reason = str(exc)
            current_size = None

        if current_size and current_size > 0:
            if current_size == last_size:
                stable_count += 1
            else:
                stable_count = 0
                last_size = current_size

            if stable_count >= stable_checks:
                if can_acquire_file_lock(filepath):
                    return True, ""
                last_reason = "文件大小已稳定，但仍无法取得独占锁"
        else:
            stable_count = 0
            last_reason = "文件为空或暂时无法读取大小"

        sleep_for = min(poll_interval, max(0, deadline - time.monotonic()))
        if sleep_for > 0:
            if stop_event:
                stop_event.wait(sleep_for)
            else:
                time.sleep(sleep_for)

    return False, f"超过 {timeout_seconds:.0f} 秒仍未完成写入：{last_reason}"


def is_file_ready(filepath: str, stable_seconds: float = 1.0) -> bool:
    ready, _ = wait_for_file_ready(
        filepath,
        timeout_seconds=max(3.0, stable_seconds * 3),
        poll_interval=stable_seconds,
        stable_checks=1,
    )
    return ready


def move_file_to_output(
    src_path: str,
    output_dir: str,
    invoice_matcher: InvoiceMatcher,
    invoice_dir_resolver=None,
    invoice_dir_planner=None,
    source_root: str = "",
    material_classifier=None,
    unmatched_action: str = "review",
    stop_event: threading.Event = None,
    dry_run: bool = False,
    operation_manifest: OperationManifest | None = None,
) -> tuple:
    """
    将文件复制到输出目录的发票号子文件夹。
    函数名保留 move_file_to_output 是为了兼容既有调用。
    返回 (success, subfolder, message, dest_path, invoice_no, status)
    """
    filename = os.path.basename(src_path)
    if stop_event and stop_event.is_set():
        return True, "", f"监控已停止，已取消处理: {filename}", src_path, "", STATUS_SKIPPED

    invoice_no = invoice_matcher.match_path(src_path, source_root=source_root)
    matched = bool(invoice_no)

    if is_path_inside(src_path, output_dir):
        return (
            True,
            "",
            f"输出目录内文件，已跳过: {filename}",
            src_path,
            invoice_no,
            STATUS_SKIPPED,
        )

    if not matched and unmatched_action == "skip":
        return (
            True,
            "",
            f"未匹配发票号，已跳过并保留原位置: {filename}",
            src_path,
            "",
            STATUS_SKIPPED,
        )

    if matched and material_classifier:
        try:
            route_classifier = getattr(
                material_classifier,
                "classify_file_for_routing",
                material_classifier.classify_file,
            )
            classification = route_classifier(src_path, invoice_no)
        except Exception as exc:
            return (
                True,
                sanitize_folder_name(invoice_no),
                f"材料识别失败，已跳过并保留原位置: {filename} | {exc}",
                src_path,
                invoice_no,
                STATUS_SKIPPED,
            )
        if classification.get("status") not in ROUTABLE_MATERIAL_STATUSES:
            material_hint = classification.get("material_name") or "未识别"
            return (
                True,
                sanitize_folder_name(invoice_no),
                f"非退税材料，已跳过并保留原位置: {filename} | 判断: {material_hint}",
                src_path,
                invoice_no,
                STATUS_SKIPPED,
            )
        if stop_event and stop_event.is_set():
            return True, "", f"监控已停止，已取消处理: {filename}", src_path, invoice_no, STATUS_SKIPPED

    subfolder_name = sanitize_folder_name(invoice_no if matched else REVIEW_SUBFOLDER)
    resolver = invoice_dir_resolver
    if dry_run and invoice_dir_planner:
        resolver = invoice_dir_planner
    dest_dir = (
        resolver(invoice_no)
        if matched and resolver
        else os.path.join(output_dir, subfolder_name)
    )
    actual_subfolder_name = os.path.basename(os.path.normpath(dest_dir)) if matched else subfolder_name

    try:
        if stop_event and stop_event.is_set():
            return True, "", f"监控已停止，已取消处理: {filename}", src_path, invoice_no, STATUS_SKIPPED
        dest_path = os.path.join(dest_dir, filename)

        dest_path, reused = resolve_copy_destination(src_path, dest_path)
        if dry_run:
            if matched:
                msg = (
                    f"预览复用: {filename} → {actual_subfolder_name}/"
                    if reused
                    else f"预览复制: {filename} → {actual_subfolder_name}/"
                )
                record_file_copy(
                    operation_manifest,
                    "copy_file",
                    src_path,
                    dest_path,
                    "planned_reuse" if reused else "planned_copy",
                    invoice_no=invoice_no,
                    subfolder=actual_subfolder_name,
                    message=msg,
                    planned=True,
                )
                return True, actual_subfolder_name, msg, dest_path, invoice_no, STATUS_PLANNED
            msg = (
                f"预览复用到人工复核: {filename} → {subfolder_name}/"
                if reused
                else f"预览复制到人工复核: {filename} → {subfolder_name}/"
            )
            record_file_copy(
                operation_manifest,
                "copy_file",
                src_path,
                dest_path,
                "planned_reuse" if reused else "planned_copy",
                invoice_no=invoice_no,
                subfolder=subfolder_name,
                message=msg,
                planned=True,
            )
            return True, subfolder_name, msg, dest_path, invoice_no, STATUS_PLANNED

        if reused:
            if matched:
                msg = f"目标已存在，已复用: {filename} → {actual_subfolder_name}/"
                record_file_copy(
                    operation_manifest,
                    "copy_file",
                    src_path,
                    dest_path,
                    "reused",
                    invoice_no=invoice_no,
                    subfolder=actual_subfolder_name,
                    message=msg,
                )
                return True, actual_subfolder_name, msg, dest_path, invoice_no, "成功"
            msg = f"未匹配发票号，人工复核中已存在相同文件: {filename} → {subfolder_name}/"
            record_file_copy(
                operation_manifest,
                "copy_file",
                src_path,
                dest_path,
                "reused",
                invoice_no=invoice_no,
                subfolder=subfolder_name,
                message=msg,
            )
            return True, subfolder_name, msg, dest_path, invoice_no, "未匹配"

        if stop_event and stop_event.is_set():
            return True, "", f"监控已停止，已取消处理: {filename}", src_path, invoice_no, STATUS_SKIPPED
        os.makedirs(os.path.dirname(dest_path), exist_ok=True)
        shutil.copy2(src_path, dest_path)
        if matched:
            msg = f"成功复制: {filename} → {actual_subfolder_name}/"
            record_file_copy(
                operation_manifest,
                "copy_file",
                src_path,
                dest_path,
                "copied",
                invoice_no=invoice_no,
                subfolder=actual_subfolder_name,
                message=msg,
            )
            return True, actual_subfolder_name, msg, dest_path, invoice_no, "成功"
        msg = f"未匹配发票号，已复制到人工复核: {filename} → {subfolder_name}/"
        record_file_copy(
            operation_manifest,
            "copy_file",
            src_path,
            dest_path,
            "copied",
            invoice_no=invoice_no,
            subfolder=subfolder_name,
            message=msg,
        )
        return (
            True,
            subfolder_name,
            msg,
            dest_path,
            invoice_no,
            "未匹配",
        )
    except Exception as e:
        return False, "", f"失败: {filename} | 原因: {e}", "", invoice_no, "失败"


# ─────────────────────────────────────────────
# 退税材料识别、判断报告、改名处理
# ─────────────────────────────────────────────

def default_material_config() -> dict:
    return {
        "folder_name": {
            "separator": "   ",
            "invoice_display": "digits_only",
        },
        "classification": {
            "auto_accept_score": 70,
            "manual_review_score": 40,
        },
        "materials": [
            {
                "id": "export_sales",
                "name": "外销",
                "required": True,
                "final_name": "外销",
                "file_name_patterns": ["外销", "proforma"],
                "pdf_text_rules": {
                    "all": ["PROFORMA INVOICE"],
                    "any": [
                        "JIAXING LAYO",
                        "Transport details",
                        "Terms of payment",
                        "Description of goods",
                    ],
                },
                "ocr_text_rules": {
                    "all": ["PROFORMA INVOICE"],
                    "any": [
                        "JIAXING LAYO",
                        "Transport details",
                        "Terms of payment",
                        "Description of goods",
                    ],
                },
                "negative_patterns": ["电子发票", "发票号码"],
            },
            {
                "id": "customs_power_of_attorney",
                "name": "报关委托书",
                "required": True,
                "final_name": "报关委托书",
                "file_name_patterns": ["报关委托书", "委托书"],
                "pdf_text_rules": {
                    "any": ["报关委托书", "代理报关委托", "报关单编号", "一般贸易"],
                    "regex_any": [r"\b\d{15,24}\b", r"\b\d{10}\b"],
                },
                "ocr_text_rules": {
                    "any": ["报关委托书", "代理报关委托", "报关单编号", "一般贸易"],
                    "regex_any": [r"\b\d{15,24}\b", r"\b\d{10}\b"],
                },
            },
            {
                "id": "customs_fee_invoice",
                "name": "报关费发票",
                "required": True,
                "final_name": "报关费发票",
                "file_name_patterns": ["报关费发票", "报关费", "报关发票"],
                "pdf_text_rules": {
                    "all": ["电子发票", "发票号码"],
                    "any": ["国内货物运输代理服务", "报关", "嘉兴新锴国际货运代理"],
                    "regex_any": [r"\b\d{2,}[A-Z]{2,}[A-Z0-9]{5,}\b"],
                },
                "ocr_text_rules": {
                    "all": ["电子发票", "发票号码"],
                    "any": ["国内货物运输代理服务", "报关", "嘉兴新锴国际货运代理"],
                    "regex_any": [r"\b\d{2,}[A-Z]{2,}[A-Z0-9]{5,}\b"],
                },
                "negative_patterns": ["国际货运代理费", "港口码头费"],
            },
            {
                "id": "bill_of_lading",
                "name": "提单",
                "required": True,
                "final_name": "提单",
                "file_name_patterns": ["提单", "海运提单", "bill of lading", "b/l"],
                "pdf_text_rules": {
                    "any": [
                        "BILL OF LADING",
                        "OCEAN BILL OF LADING",
                        "OCEAN OR COMBINED TRANSPORT BILL OF LADING",
                        "SEA WAYBILL",
                        "BILLOFLADING",
                        "Bill of Lading No.",
                        "Number of Original B/L",
                        "TELEX RELEASE",
                        "SHIPPER",
                        "CONSIGNEE",
                        "NOTIFY PARTY",
                        "FREIGHT COLLECT",
                        "SHIPPED ON BOARD",
                        "SHIPPER'S LOAD COUNT",
                    ],
                    "regex_any": [r"[A-Z]{4}\d{7}", r"\d+\s*CARTONS", r"\d+\.\d+KGS"],
                },
                "ocr_text_rules": {
                    "any": [
                        "BILL OF LADING",
                        "OCEAN BILL OF LADING",
                        "OCEAN OR COMBINED TRANSPORT BILL OF LADING",
                        "SEA WAYBILL",
                        "BILLOFLADING",
                        "Bill of Lading No.",
                        "Number of Original B/L",
                        "TELEX RELEASE",
                        "SHIPPER",
                        "CONSIGNEE",
                        "NOTIFY PARTY",
                        "FREIGHT COLLECT",
                        "SHIPPED ON BOARD",
                        "SHIPPER'S LOAD COUNT",
                    ],
                    "regex_any": [r"[A-Z]{4}\d{7}", r"\d+\s*CARTONS", r"\d+\.\d+KGS"],
                },
            },
            {
                "id": "packing_list",
                "name": "装箱单",
                "required": True,
                "final_name": "装箱单",
                "file_name_patterns": ["装箱单", "箱单", "packing list"],
                "pdf_text_rules": {
                    "any": [
                        "PACKING LIST",
                        "装箱单",
                        "CONTAINER LOAD PLAN",
                        "Container No.",
                        "Seal No.",
                        "Port of Loading",
                        "Port of Discharge",
                        "Place of Delivery",
                        "Packing Date",
                        "Total Packages",
                        "Total Cargo Wt",
                        "Total Meas",
                        "CARTONS",
                        "G.W.",
                        "N.W.",
                        "MEAS",
                        "SHANGHAI",
                        "GDANSK",
                        "40HC",
                        "40HQ",
                        "收货作业记录单",
                        "收货凭证",
                        "进仓编号",
                        "来自地点",
                        "送货车号",
                        "客户名称",
                        "司机号码",
                        "登记件数",
                        "包装类型",
                        "总件数",
                        "总体积",
                        "总重量",
                        "残损记录",
                        "实收件数",
                        "分票类型",
                        "车辆类型",
                        "换单员",
                        "收货员",
                        "库位",
                        "唛头",
                        "件数",
                    ],
                    "regex_any": [
                        r"\b[A-Z]{4}\d{7}\b",
                        r"\b[A-Z]{1,4}\d{5,10}\b",
                        r"\b40H[QC]\b",
                        r"\bJC\d{6,10}\b",
                        r"\bL\d{7,}[A-Z0-9]*\b",
                        r"[\u4e00-\u9fa5][A-Z][A-Z0-9]{5}\b",
                    ],
                },
                "ocr_text_rules": {
                    "any": [
                        "PACKING LIST",
                        "装箱单",
                        "CONTAINER LOAD PLAN",
                        "Container No.",
                        "Seal No.",
                        "Port of Loading",
                        "Port of Discharge",
                        "Place of Delivery",
                        "Packing Date",
                        "Total Packages",
                        "Total Cargo Wt",
                        "Total Meas",
                        "CARTONS",
                        "G.W.",
                        "N.W.",
                        "MEAS",
                        "SHANGHAI",
                        "GDANSK",
                        "40HC",
                        "40HQ",
                        "收货作业记录单",
                        "收货凭证",
                        "进仓编号",
                        "来自地点",
                        "送货车号",
                        "客户名称",
                        "司机号码",
                        "登记件数",
                        "包装类型",
                        "总件数",
                        "总体积",
                        "总重量",
                        "残损记录",
                        "实收件数",
                        "分票类型",
                        "车辆类型",
                        "换单员",
                        "收货员",
                        "库位",
                        "唛头",
                        "件数",
                    ],
                    "regex_any": [
                        r"\b[A-Z]{4}\d{7}\b",
                        r"\b[A-Z]{1,4}\d{5,10}\b",
                        r"\b40H[QC]\b",
                        r"\bJC\d{6,10}\b",
                        r"\bL\d{7,}[A-Z0-9]*\b",
                        r"[\u4e00-\u9fa5][A-Z][A-Z0-9]{5}\b",
                    ],
                },
                "negative_patterns": [
                    "OCEAN OR COMBINED TRANSPORT BILL OF LADING",
                    "BILLOFLADING",
                    "FREIGHT COLLECT",
                    "SHIPPED ON BOARD",
                    "TELEX RELEASE",
                    "ORIGINAL B/L",
                    "SHIPPER'S LOAD",
                ],
            },
            {
                "id": "freight_invoice",
                "name": "运费发票",
                "required": True,
                "final_name": "运费发票",
                "file_name_patterns": ["运费发票", "运费"],
                "pdf_text_rules": {
                    "all": ["电子发票", "发票号码"],
                    "any": ["国际货运代理费", "港口码头费", "汇利达欧海国际货运代理"],
                    "regex_any": [r"\b[A-Z]{5,8}\d{6,12}\b"],
                },
                "ocr_text_rules": {
                    "all": ["电子发票", "发票号码"],
                    "any": ["国际货运代理费", "港口码头费", "汇利达欧海国际货运代理"],
                    "regex_any": [r"\b[A-Z]{5,8}\d{6,12}\b"],
                },
                "negative_patterns": ["国内货物运输代理服务", "报关费"],
            },
            {
                "id": "purchase_contract",
                "name": "采购合同",
                "required": False,
                "final_name": "采购合同",
                "file_name_only": True,
                "file_name_patterns": ["采购合同", "合同"],
            },
        ],
    }


def ensure_material_config(path: Path = DEFAULT_MATERIAL_CONFIG) -> dict:
    if not path.exists():
        data = default_material_config()
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return data

    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)

    changed = False

    def add_values(target: list, values: list[str]) -> bool:
        local_changed = False
        for value in values:
            if value not in target:
                target.append(value)
                local_changed = True
        return local_changed

    def ensure_rules(material: dict, rules_key: str, any_values: list[str] = None,
                     regex_values: list[str] = None, all_values: list[str] = None) -> bool:
        rules = material.setdefault(rules_key, {})
        local_changed = False
        if all_values:
            local_changed = add_values(rules.setdefault("all", []), all_values) or local_changed
        if any_values:
            local_changed = add_values(rules.setdefault("any", []), any_values) or local_changed
        if regex_values:
            local_changed = add_values(rules.setdefault("regex_any", []), regex_values) or local_changed
        return local_changed

    export_sales_any = [
        "JIAXING LAYO",
        "Transport details",
        "Terms of payment",
        "Description of goods",
    ]
    customs_any = ["报关委托书", "代理报关委托", "报关单编号", "一般贸易"]
    customs_regex = [r"\b\d{15,24}\b", r"\b\d{10}\b"]
    customs_fee_any = ["国内货物运输代理服务", "报关", "嘉兴新锴国际货运代理"]
    customs_fee_regex = [r"\b\d{2,}[A-Z]{2,}[A-Z0-9]{5,}\b"]
    bill_any = [
        "BILL OF LADING",
        "OCEAN BILL OF LADING",
        "OCEAN OR COMBINED TRANSPORT BILL OF LADING",
        "SEA WAYBILL",
        "BILLOFLADING",
        "Bill of Lading No.",
        "Number of Original B/L",
        "TELEX RELEASE",
        "SHIPPER",
        "CONSIGNEE",
        "NOTIFY PARTY",
        "FREIGHT COLLECT",
        "SHIPPED ON BOARD",
        "SHIPPER'S LOAD COUNT",
    ]
    bill_regex = [r"[A-Z]{4}\d{7}", r"\d+\s*CARTONS", r"\d+\.\d+KGS"]
    packing_any = [
        "PACKING LIST",
        "装箱单",
        "CONTAINER LOAD PLAN",
        "Container No.",
        "Seal No.",
        "Port of Loading",
        "Port of Discharge",
        "Place of Delivery",
        "Packing Date",
        "Total Packages",
        "Total Cargo Wt",
        "Total Meas",
        "CARTONS",
        "G.W.",
        "N.W.",
        "MEAS",
        "SHANGHAI",
        "GDANSK",
        "40HC",
        "40HQ",
        "收货作业记录单",
        "收货凭证",
        "进仓编号",
        "来自地点",
        "送货车号",
        "客户名称",
        "司机号码",
        "登记件数",
        "包装类型",
        "总件数",
        "总体积",
        "总重量",
        "残损记录",
        "实收件数",
        "分票类型",
        "车辆类型",
        "换单员",
        "收货员",
        "库位",
        "唛头",
        "件数",
    ]
    packing_regex = [
        r"\b[A-Z]{4}\d{7}\b",
        r"\b[A-Z]{1,4}\d{5,10}\b",
        r"\b40H[QC]\b",
        r"\bJC\d{6,10}\b",
        r"\bL\d{7,}[A-Z0-9]*\b",
        r"[\u4e00-\u9fa5][A-Z][A-Z0-9]{5}\b",
    ]
    packing_negative = [
        "OCEAN OR COMBINED TRANSPORT BILL OF LADING",
        "BILLOFLADING",
        "FREIGHT COLLECT",
        "SHIPPED ON BOARD",
        "TELEX RELEASE",
        "ORIGINAL B/L",
        "SHIPPER'S LOAD",
    ]
    freight_any = ["国际货运代理费", "港口码头费", "汇利达欧海国际货运代理"]
    freight_regex = [r"\b[A-Z]{5,8}\d{6,12}\b"]

    for material in data.get("materials", []):
        if material.get("id") == "purchase_contract":
            patterns = material.setdefault("file_name_patterns", [])
            if "合同" not in patterns:
                patterns.append("合同")
                changed = True
        if material.get("id") == "export_sales":
            changed = ensure_rules(
                material,
                "ocr_text_rules",
                any_values=export_sales_any,
                all_values=["PROFORMA INVOICE"],
            ) or changed
        if material.get("id") == "customs_power_of_attorney":
            for rules_key in ["pdf_text_rules", "ocr_text_rules"]:
                changed = ensure_rules(
                    material,
                    rules_key,
                    any_values=customs_any,
                    regex_values=customs_regex,
                ) or changed
        if material.get("id") == "customs_fee_invoice":
            for rules_key in ["pdf_text_rules", "ocr_text_rules"]:
                changed = ensure_rules(
                    material,
                    rules_key,
                    any_values=customs_fee_any,
                    regex_values=customs_fee_regex,
                    all_values=["电子发票", "发票号码"],
                ) or changed
        if material.get("id") == "bill_of_lading":
            patterns = material.setdefault("file_name_patterns", [])
            for pattern in ["海运提单"]:
                if pattern not in patterns:
                    patterns.append(pattern)
                    changed = True

            for rules_key in ["pdf_text_rules", "ocr_text_rules"]:
                changed = ensure_rules(
                    material,
                    rules_key,
                    any_values=bill_any,
                    regex_values=bill_regex,
                ) or changed
        if material.get("id") == "packing_list":
            changed = add_values(material.setdefault("negative_patterns", []), packing_negative) or changed
            for rules_key in ["pdf_text_rules", "ocr_text_rules"]:
                changed = ensure_rules(
                    material,
                    rules_key,
                    any_values=packing_any,
                    regex_values=packing_regex,
                ) or changed
        if material.get("id") == "freight_invoice":
            for rules_key in ["pdf_text_rules", "ocr_text_rules"]:
                changed = ensure_rules(
                    material,
                    rules_key,
                    any_values=freight_any,
                    regex_values=freight_regex,
                    all_values=["电子发票", "发票号码"],
                ) or changed
    if changed:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    return data


def keyword_match(pattern: str, text: str) -> bool:
    if not pattern:
        return False
    if pattern.startswith("re:"):
        return re.search(pattern[3:], text, re.IGNORECASE) is not None
    return normalize_text_for_match(pattern) in normalize_text_for_match(text)


def pdf_text(filepath: str, max_pages: int = 2) -> str:
    if Path(filepath).suffix.lower() != ".pdf":
        return ""
    try:
        import fitz
    except Exception:
        return ""

    try:
        doc = fitz.open(filepath)
        try:
            chunks = []
            for page in doc[:max_pages]:
                chunks.append(page.get_text("text"))
            return "\n".join(chunks).strip()
        finally:
            doc.close()
    except Exception as exc:
        logging.error(f"PDF 文本提取失败: {filepath} | {exc}")
        return ""


def rapid_ocr_text(filepath: str) -> str:
    try:
        from rapidocr_onnxruntime import RapidOCR
    except Exception:
        try:
            from rapidocr import RapidOCR
        except Exception:
            return ""

    try:
        engine = rapid_ocr_text._engine
    except AttributeError:
        engine = RapidOCR()
        rapid_ocr_text._engine = engine

    try:
        result, _ = engine(filepath)
    except Exception as exc:
        logging.error(f"OCR 识别失败: {filepath} | {exc}")
        return ""
    if not result:
        return ""
    return "\n".join(str(item[1]) for item in result if len(item) >= 2)


def _save_ocr_variant(image, variant_paths: list[str]) -> None:
    with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
        tmp_path = tmp.name
    try:
        image.save(tmp_path, format="PNG")
        variant_paths.append(tmp_path)
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def _resize_image_for_ocr(image):
    width, height = image.size
    long_side = max(width, height)
    short_side = max(1, min(width, height))
    scale = 1.0
    if short_side < 1200:
        scale = min(3.0, 1200 / short_side)
    elif long_side < 2200:
        scale = min(2.0, 2200 / long_side)
    if scale <= 1.05:
        return image.copy()

    resampling = getattr(getattr(image, "Resampling", None), "LANCZOS", None)
    if resampling is None:
        try:
            from PIL import Image
            resampling = getattr(Image, "Resampling", Image).LANCZOS
        except Exception:
            resampling = 1
    return image.resize((int(width * scale), int(height * scale)), resampling)


def _create_preprocessed_image_variants(filepath: str) -> list[str]:
    try:
        from PIL import Image, ImageFilter, ImageOps
    except Exception:
        return []

    variant_paths = []
    try:
        with Image.open(filepath) as original:
            image = ImageOps.exif_transpose(original).convert("RGB")
            gray = ImageOps.grayscale(image)
            gray = ImageOps.autocontrast(gray, cutoff=1)
            gray = _resize_image_for_ocr(gray)

            _save_ocr_variant(gray, variant_paths)
            _save_ocr_variant(gray.filter(ImageFilter.SHARPEN), variant_paths)

            histogram = gray.histogram()
            total = sum(histogram) or 1
            mean = sum(index * count for index, count in enumerate(histogram)) / total
            threshold = min(205, max(145, int(mean * 0.9)))
            binary = gray.point(lambda pixel: 255 if pixel > threshold else 0).convert("L")
            _save_ocr_variant(binary, variant_paths)

            width, height = gray.size
            if max(width, height) / max(1, min(width, height)) >= 1.25:
                _save_ocr_variant(gray.rotate(90, expand=True), variant_paths)
                _save_ocr_variant(gray.rotate(270, expand=True), variant_paths)
    except Exception as exc:
        logging.error(f"图片 OCR 预处理失败: {filepath} | {exc}")
        for path in variant_paths:
            try:
                os.remove(path)
            except OSError:
                pass
        return []
    return variant_paths


def _merge_ocr_texts(texts: list[str]) -> str:
    merged_lines = []
    seen = set()
    for text in texts:
        for line in str(text or "").splitlines():
            stripped = line.strip()
            if not stripped:
                continue
            key = normalize_text_for_match(stripped)
            if key in seen:
                continue
            seen.add(key)
            merged_lines.append(stripped)
    return "\n".join(merged_lines).strip()


def _ocr_image_text(filepath: str) -> str:
    texts = []
    temp_paths = []
    try:
        raw_text = rapid_ocr_text(filepath)
        if raw_text:
            texts.append(raw_text)
        temp_paths = _create_preprocessed_image_variants(filepath)
        for path in temp_paths:
            text = rapid_ocr_text(path)
            if text:
                texts.append(text)
    finally:
        for path in temp_paths:
            try:
                os.remove(path)
            except OSError:
                pass
    return _merge_ocr_texts(texts)


def ocr_text(filepath: str, max_pdf_pages: int = 1) -> str:
    suffix = Path(filepath).suffix.lower()
    if suffix in IMAGE_OCR_SUFFIXES:
        return _ocr_image_text(filepath)
    if suffix != ".pdf":
        return ""

    try:
        import fitz
    except Exception:
        return ""

    texts = []
    try:
        doc = fitz.open(filepath)
        try:
            for page in doc[:max_pdf_pages]:
                pix = page.get_pixmap(matrix=fitz.Matrix(2, 2), alpha=False)
                with tempfile.NamedTemporaryFile(suffix=".png", delete=False) as tmp:
                    tmp_path = tmp.name
                try:
                    pix.save(tmp_path)
                    text = _ocr_image_text(tmp_path)
                    if text:
                        texts.append(text)
                finally:
                    try:
                        os.remove(tmp_path)
                    except OSError:
                        pass
        finally:
            doc.close()
    except Exception as exc:
        logging.error(f"PDF OCR 渲染失败: {filepath} | {exc}")
    return "\n".join(texts).strip()


def extract_declaration_number(text: str) -> str:
    if not text:
        return ""
    candidates = re.findall(r"\b\d{15,24}\b", normalize_document_text(text))
    return candidates[0] if candidates else ""


def extract_amounts(text: str) -> list[float]:
    amounts = []
    normalized_text = normalize_document_text(text)
    for match in re.findall(r"(?:USD|RMB|CNY|¥|\$)?\s*-?\d[\d,]*(?:\.\d{1,4})?", normalized_text, re.IGNORECASE):
        amount = normalize_amount(match)
        if amount is not None:
            amounts.append(amount)
    return amounts


def extract_dates(text: str) -> list[str]:
    values = []
    if not text:
        return values
    text = normalize_document_text(text)

    patterns = [
        r"\b\d{4}[-/.]\d{1,2}[-/.]\d{1,2}\b",
        r"\d{4}年\d{1,2}月\d{1,2}日",
        r"\b(?:JANUARY|JAN|FEBRUARY|FEB|MARCH|MAR|APRIL|APR|MAY|JUNE|JUN|JULY|JUL|AUGUST|AUG|SEPTEMBER|SEPT|SEP|OCTOBER|OCT|NOVEMBER|NOV|DECEMBER|DEC)\.?\s*,?\s*\d{1,2}\s*,?\s*\d{4}\b",
    ]
    for pattern in patterns:
        for raw in re.findall(pattern, text, re.IGNORECASE):
            normalized = normalize_date_value(raw)
            if normalized and normalized not in values:
                values.append(normalized)
    return values


def extract_document_fields(material_id: str, text: str, filename: str) -> dict:
    raw_haystack = f"{filename}\n{text or ''}"
    haystack = normalize_document_text(raw_haystack)
    content_text = normalize_document_text(text or "")
    invoice_candidates = re.findall(r"\b[A-Z]{1,4}\d{6,12}\b", haystack, re.IGNORECASE)
    invoice_candidates.extend(re.findall(r"\b\d{6,12}\b", haystack))
    fields = {
        "invoice_numbers": sorted(set(invoice_candidates)),
        "customs_declaration_numbers": [],
        "bl_numbers": [],
        "booking_refs": [],
        "container_numbers": [],
        "seal_numbers": [],
        "amounts": extract_amounts(content_text),
        "dates": extract_dates(content_text),
        "countries": [],
        "raw_text_sample": re.sub(r"\s+", " ", raw_haystack).strip()[:500],
        "normalized_text_sample": haystack[:500],
    }

    upper = content_text.upper()
    for key, country in COUNTRY_MAP.items():
        if key in upper and country not in fields["countries"]:
            fields["countries"].append(country)

    container_numbers = set(re.findall(r"\b[A-Z]{4}\d{7}\b", upper))
    fields["container_numbers"] = sorted(container_numbers)

    # 提单号通常是 5-8 位字母 + 6-12 位数字，避免与 4 字母+7数字的箱号混淆。
    bl_numbers = {
        value for value in re.findall(r"\b[A-Z]{5,8}\d{6,12}\b", upper)
        if value not in container_numbers
    }
    if material_id in {"bill_of_lading", "freight_invoice"}:
        fields["bl_numbers"] = sorted(bl_numbers)

    booking_refs = set(re.findall(r"\b\d{2,}[A-Z]{2,}[A-Z0-9]{5,}\b", upper))
    if material_id in {"customs_power_of_attorney", "customs_fee_invoice", "packing_list"}:
        fields["booking_refs"] = sorted(booking_refs)

    long_numbers = re.findall(r"\b\d{15,24}\b", upper)
    if material_id == "customs_power_of_attorney":
        fields["customs_declaration_numbers"] = long_numbers[:3]

    seal_candidates = {
        value for value in re.findall(r"\b[A-Z]{1,4}\d{5,10}\b", upper)
        if value not in container_numbers and value not in bl_numbers
    }
    if material_id in {"bill_of_lading", "packing_list"}:
        fields["seal_numbers"] = sorted(seal_candidates)

    return fields


def any_overlap(left: list, right: list) -> bool:
    return bool(set(left or []) & set(right or []))


def any_identifier_overlap(left: list, right: list, min_prefix_length: int = 10, max_length_gap: int = 2) -> bool:
    left_values = [normalize_identifier(value) for value in left or []]
    right_values = [normalize_identifier(value) for value in right or []]
    for left_value in left_values:
        if not left_value:
            continue
        for right_value in right_values:
            if not right_value:
                continue
            if left_value == right_value:
                return True
            shorter, longer = sorted((left_value, right_value), key=len)
            if (
                len(shorter) >= min_prefix_length
                and len(longer) - len(shorter) <= max_length_gap
                and longer.startswith(shorter)
            ):
                return True
    return False


class MaterialClassifier:
    def __init__(self, config: dict):
        self.config = config
        self.materials = config.get("materials", [])
        self.auto_score = int(config.get("classification", {}).get("auto_accept_score", 70))
        self.manual_score = int(config.get("classification", {}).get("manual_review_score", 40))

    def required_names(self) -> list[str]:
        return [m["name"] for m in self.materials if m.get("required")]

    def classify_file_for_routing(self, filepath: str, invoice_no: str) -> dict:
        """Fast path used while files enter the watch folder.

        Detailed closed-loop checks still call classify_file(), which can use
        PDF extraction and OCR. Routing only needs to know whether a file looks
        like a tax material, so a clear material keyword in the file name is
        enough to keep the UI responsive during large drops.
        """
        filename = os.path.basename(filepath)
        stripped_name = strip_invoice_prefix(filename, invoice_no)
        best = None
        for material in self.materials:
            score, reasons = self._score_file_name(material, stripped_name)
            if best is None or score > best["score"]:
                best = {
                    "material": material,
                    "score": score,
                    "reasons": reasons,
                }

        if best and best["score"] >= self.manual_score:
            material = best["material"]
            return {
                "filename": filename,
                "path": filepath,
                "stripped_name": stripped_name,
                "material": material,
                "material_id": material["id"],
                "material_name": material["name"],
                "required": bool(material.get("required")),
                "score": best["score"],
                "status": "matched" if material.get("file_name_only") else "manual_review",
                "reasons": best["reasons"],
                "text": "",
                "normalized_text": "",
            }

        return self.classify_file(filepath, invoice_no)

    def classify_file(self, filepath: str, invoice_no: str) -> dict:
        filename = os.path.basename(filepath)
        stripped_name = strip_invoice_prefix(filename, invoice_no)
        for material in self.materials:
            if not material.get("file_name_only"):
                continue
            score, reasons = self._score_file_name(material, stripped_name)
            if score >= self.manual_score:
                return {
                    "filename": filename,
                    "path": filepath,
                    "stripped_name": stripped_name,
                    "material": material,
                    "material_id": material["id"],
                    "material_name": material["name"],
                    "required": bool(material.get("required")),
                    "score": score,
                    "status": "matched" if score >= self.manual_score else "unrecognized",
                    "reasons": reasons,
                    "text": "",
                    "normalized_text": "",
                }

        text_cache = {"pdf": None, "ocr": None}
        best = None
        all_scores = []

        for material in self.materials:
            score, reasons = self._score_material(material, filepath, stripped_name, text_cache)
            if score:
                all_scores.append({
                    "material_id": material["id"],
                    "material_name": material["name"],
                    "score": score,
                    "reasons": reasons,
                })
            if best is None or score > best["score"]:
                best = {
                    "material": material,
                    "score": score,
                    "reasons": reasons,
                }

        if not best or best["score"] < self.manual_score:
            status = "unrecognized"
            material = None
        else:
            tied = [
                item for item in all_scores
                if item["material_id"] != best["material"]["id"] and abs(item["score"] - best["score"]) <= 10
            ]
            has_content_match = any(
                reason.startswith(("PDF", "OCR"))
                for reason in best["reasons"]
            )
            if best["material"].get("file_name_only"):
                status = "matched" if best["score"] >= self.manual_score else "unrecognized"
            elif tied or best["score"] < self.auto_score or not has_content_match:
                status = "manual_review"
            else:
                status = "matched"
            material = best["material"]

        text = ""
        if text_cache["pdf"]:
            text = text_cache["pdf"]
        elif text_cache["ocr"]:
            text = text_cache["ocr"]

        return {
            "filename": filename,
            "path": filepath,
            "stripped_name": stripped_name,
            "material": material,
            "material_id": material["id"] if material else "",
            "material_name": material["name"] if material else "",
            "required": bool(material and material.get("required")),
            "score": best["score"] if best else 0,
            "status": status,
            "reasons": best["reasons"] if best else [],
            "text": text,
            "normalized_text": normalize_document_text(text),
        }

    def _score_material(self, material: dict, filepath: str, stripped_name: str, text_cache: dict) -> tuple[int, list[str]]:
        score, reasons = self._score_file_name(material, stripped_name)

        if material.get("file_name_only"):
            return score, reasons

        pdf_rules = material.get("pdf_text_rules") or {}
        ocr_rules = material.get("ocr_text_rules") or {}

        if pdf_rules:
            if text_cache["pdf"] is None:
                text_cache["pdf"] = pdf_text(filepath)
            pdf_score, pdf_reasons = self._score_text_rules(pdf_rules, text_cache["pdf"], "PDF")
            score += pdf_score
            reasons.extend(pdf_reasons)

        if ocr_rules and Path(filepath).suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf"}:
            if text_cache["ocr"] is None:
                text_cache["ocr"] = ocr_text(filepath)
            ocr_score, ocr_reasons = self._score_text_rules(ocr_rules, text_cache["ocr"], "OCR")
            score += ocr_score
            reasons.extend(ocr_reasons)

        text_for_negative = "\n".join(filter(None, [text_cache["pdf"] or "", text_cache["ocr"] or "", stripped_name]))
        negative_hit = False
        for pattern in material.get("negative_patterns", []):
            if keyword_match(pattern, text_for_negative):
                negative_hit = True
                score -= 50
                reasons.append(f"排除词命中:{pattern}")

        if material.get("id") == "packing_list" and not negative_hit:
            fallback_score, fallback_reasons = self._score_packing_list_ocr_fallback(
                score,
                reasons,
                stripped_name,
                text_cache["pdf"] or "",
                text_cache["ocr"] or "",
            )
            score += fallback_score
            reasons.extend(fallback_reasons)

        return max(score, 0), reasons

    @staticmethod
    def _score_packing_list_ocr_fallback(
        score: int,
        reasons: list[str],
        stripped_name: str,
        pdf_text_value: str,
        ocr_text_value: str,
    ) -> tuple[int, list[str]]:
        if score < 25:
            return 0, []
        if not (keyword_match("装箱单", stripped_name) or keyword_match("箱单", stripped_name)):
            return 0, []

        source = "OCR" if ocr_text_value else "PDF"
        text = normalize_document_text(ocr_text_value or pdf_text_value)
        if len(text) < 80:
            return 0, []

        warehouse_signals = [
            "收货", "进仓", "件数", "体积", "重量", "唛头", "库位", "车号",
            "客户名称", "司机", "包装", "货物", "凭证", "作业记录",
        ]
        matched_signals = [signal for signal in warehouse_signals if keyword_match(signal, text)]
        if len(matched_signals) >= 2:
            return 30, [f"{source}仓库装箱字段组合命中:{','.join(matched_signals[:4])}"]

        has_content_match = any(reason.startswith(("PDF", "OCR")) for reason in reasons)
        if not has_content_match and score < 70 and len(text) >= 120:
            return 30, [f"{source}文本可用:装箱单文件名兜底"]
        return 0, []

    @staticmethod
    def _score_file_name(material: dict, stripped_name: str) -> tuple[int, list[str]]:
        score = 0
        reasons = []
        for pattern in material.get("file_name_patterns", []):
            if keyword_match(pattern, stripped_name):
                score += 60 if material.get("file_name_only") else 25
                reasons.append(f"文件名命中:{pattern}")
                break
        final_name = material.get("final_name") or material.get("name")
        if final_name and keyword_match(final_name, stripped_name):
            score += 20 if material.get("file_name_only") else 15
            reasons.append(f"文件名标准名命中:{final_name}")
        return score, reasons

    @staticmethod
    def _score_text_rules(rules: dict, text: str, source: str) -> tuple[int, list[str]]:
        if not text:
            return 0, []
        normalized_text = normalize_document_text(text)
        score = 0
        reasons = []
        match_count = 0

        all_patterns = rules.get("all", [])
        if all_patterns and all(keyword_match(pattern, normalized_text) for pattern in all_patterns):
            score += 50
            match_count += len(all_patterns)
            reasons.append(f"{source}全部命中:{','.join(all_patterns)}")

        for pattern in rules.get("any", []):
            if keyword_match(pattern, normalized_text):
                score += 20
                match_count += 1
                reasons.append(f"{source}命中:{pattern}")

        for pattern in rules.get("regex_any", []):
            if re.search(pattern, normalized_text, re.IGNORECASE):
                score += 20
                match_count += 1
                reasons.append(f"{source}正则命中:{pattern}")

        if match_count >= 2:
            score += 20
            reasons.append(f"{source}多字段组合命中")

        return score, reasons


def evaluate_closed_loop(invoice_no: str, ledger_record: dict, classifications: list[dict], declaration_no: str) -> dict:
    docs = {
        item["material_id"]: item
        for item in classifications
        if item.get("status") == "matched" and item.get("material_id")
    }

    score = 0
    evidence = []
    conflicts = []
    unverified = []

    sales = docs.get("export_sales")
    customs = docs.get("customs_power_of_attorney")
    bill = docs.get("bill_of_lading")
    packing = docs.get("packing_list")
    freight = docs.get("freight_invoice")
    customs_fee = docs.get("customs_fee_invoice")

    ledger_invoice_ok = bool(ledger_record and invoice_equivalent(ledger_record.get("发票号", ""), invoice_no))
    sales_invoice_ok = False
    sales_invoice_values = []
    if sales:
        sales_invoice_values = sales.get("fields", {}).get("invoice_numbers", [])
        sales_invoice_ok = any(invoice_equivalent(value, invoice_no) for value in sales_invoice_values)
    if ledger_invoice_ok or sales_invoice_ok:
        score += 30
        evidence.append("外销发票号/台账发票号与当前发票号一致(+30)")
    if sales and not sales_invoice_ok:
        if sales_invoice_values:
            conflicts.append("外销发票号与当前发票号不一致")
        else:
            conflicts.append("外销发票未提取到匹配的发票号")

    if ledger_record and not ledger_invoice_ok:
        conflicts.append("台账发票号与当前发票号不一致")
    if not ledger_record:
        unverified.extend(["台账发票号未核验", "台账金额未核验", "台账目的国未核验", "台账出运日期未核验"])

    customs_decl_ok = False
    if customs:
        customs_numbers = customs.get("fields", {}).get("customs_declaration_numbers", [])
        customs_decl_ok = bool(
            declaration_no and declaration_no in customs_numbers
        )
        if customs_decl_ok:
            score += 30
            evidence.append("报关委托书报关单号与当前报关号一致(+30)")
        elif declaration_no and customs_numbers:
            conflicts.append("报关委托书报关单号与当前报关号不一致")

    bill_freight_ok = False
    if bill and freight:
        bill_freight_ok = any_overlap(
            bill.get("fields", {}).get("bl_numbers", []),
            freight.get("fields", {}).get("bl_numbers", []),
        )
        if bill_freight_ok:
            score += 25
            evidence.append("提单号连接提单与运费发票(+25)")

    booking_customs_fee_ok = False
    if customs and customs_fee:
        booking_customs_fee_ok = any_identifier_overlap(
            customs.get("fields", {}).get("booking_refs", []),
            customs_fee.get("fields", {}).get("booking_refs", []),
        )
        if booking_customs_fee_ok:
            score += 25
            evidence.append("提运参考号连接报关委托书与报关费发票(+25)")

    container_ok = False
    if bill and packing:
        container_ok = any_overlap(
            bill.get("fields", {}).get("container_numbers", []),
            packing.get("fields", {}).get("container_numbers", []),
        )
        if container_ok:
            score += 20
            evidence.append("箱号连接提单与装箱单(+20)")

        seal_ok = any_overlap(
            bill.get("fields", {}).get("seal_numbers", []),
            packing.get("fields", {}).get("seal_numbers", []),
        )
        if seal_ok:
            score += 20
            evidence.append("封号连接提单与装箱单(+20)")
    else:
        seal_ok = False

    ledger_amount_ok = False
    ledger_country_ok = False
    ledger_date_ok = False

    if ledger_record and sales:
        ledger_amount = ledger_record.get("金额")
        sales_amounts = sales.get("fields", {}).get("amounts", [])
        if ledger_amount is not None and any(amounts_equal(ledger_amount, amount) for amount in sales_amounts):
            ledger_amount_ok = True
            score += 15
            evidence.append("台账金额与外销发票金额一致(+15)")
        elif ledger_amount is not None and sales_amounts:
            conflicts.append("台账金额与外销发票金额未匹配")
        elif ledger_amount in (None, ""):
            unverified.append("台账金额未核验")

        ledger_country = ledger_record.get("目的国", "")
        sales_countries = sales.get("fields", {}).get("countries", [])
        if ledger_country and ledger_country in sales_countries:
            ledger_country_ok = True
            score += 10
            evidence.append("台账目的国与外销发票国家一致(+10)")
        elif ledger_country and sales_countries:
            conflicts.append("台账目的国与外销发票国家不一致")
        elif not ledger_country:
            unverified.append("台账目的国未核验")

    if ledger_record and bill:
        ledger_date = ledger_record.get("出运日期", "")
        bill_dates = bill.get("fields", {}).get("dates", [])
        if ledger_date and any(dates_close(ledger_date, item) for item in bill_dates):
            ledger_date_ok = True
            score += 15
            evidence.append("台账出运日期与提单日期接近(+15)")
        elif ledger_date and bill_dates:
            conflicts.append("台账出运日期与提单日期相差过大")
        elif not ledger_date:
            unverified.append("台账出运日期未核验")

    strong_combo = any([
        sales_invoice_ok and ledger_amount_ok and ledger_country_ok,
        customs_decl_ok and booking_customs_fee_ok,
        bill_freight_ok and container_ok and seal_ok,
        sales_invoice_ok and ledger_date_ok and (bill_freight_ok or booking_customs_fee_ok),
    ])

    if score >= 90 and strong_combo and not conflicts:
        status = "strong_pass"
        status_label = "强闭环通过"
    elif score >= 70 and strong_combo and not conflicts:
        status = "basic_pass"
        status_label = "基本闭环通过"
    elif score >= 50:
        status = "review"
        status_label = "待复核"
    else:
        status = "failed"
        status_label = "不通过"

    return {
        "status": status,
        "status_label": status_label,
        "score": score,
        "evidence": evidence,
        "conflicts": conflicts,
        "unverified": sorted(set(unverified)),
        "strong_combo": strong_combo,
    }


class TaxRefundProcessor:
    def __init__(
        self,
        output_dir: str,
        config_path: Path = DEFAULT_MATERIAL_CONFIG,
        ledger_records: dict = None,
        dry_run: bool = False,
        operation_manifest: OperationManifest | None = None,
        copy_first_rename: bool = True,
    ):
        self.output_dir = output_dir
        self.config_path = config_path
        self.config = ensure_material_config(config_path)
        self.classifier = MaterialClassifier(self.config)
        self.ledger_records = ledger_records or {}
        self.dry_run = bool(dry_run)
        self.operation_manifest = operation_manifest
        self.copy_first_rename = bool(copy_first_rename)
        self.lock = threading.Lock()
        self.declaration_path = os.path.join(output_dir, "declaration_overrides.json")
        self.declarations = self._load_declarations()

    def plan_target_folder(self, invoice_no: str) -> str:
        return self._target_folder_for_invoice(invoice_no, allow_copy_merge=False)

    def resolve_target_folder(self, invoice_no: str, allow_copy_merge: bool = True) -> str:
        if not self.dry_run:
            os.makedirs(self.output_dir, exist_ok=True)
        return self._target_folder_for_invoice(invoice_no, allow_copy_merge=allow_copy_merge)

    def _target_folder_for_invoice(self, invoice_no: str, allow_copy_merge: bool = True) -> str:
        original = os.path.join(self.output_dir, sanitize_folder_name(invoice_no))

        final_like_folder = self._find_invoice_suffix_folder(invoice_no)
        if final_like_folder:
            return final_like_folder

        digits = invoice_digits(invoice_no)
        if digits:
            digits_folder = os.path.join(self.output_dir, sanitize_folder_name(digits))
            if os.path.isdir(digits_folder):
                self._copy_merge_original_folder(original, digits_folder, allow_copy_merge)
                return digits_folder

        prefixed_folder = self._find_prefixed_invoice_folder(invoice_no)
        if prefixed_folder:
            self._copy_merge_original_folder(original, prefixed_folder, allow_copy_merge)
            return prefixed_folder

        if os.path.isdir(original):
            return original

        return original

    def _copy_merge_original_folder(self, original: str, target: str, allow_copy_merge: bool) -> None:
        if (
            allow_copy_merge
            and os.path.isdir(original)
            and os.path.isdir(target)
            and not paths_equal(original, target)
        ):
            copied = copy_folder_contents(
                original,
                target,
                operation_manifest=self.operation_manifest,
                action="copy_merge_file",
                dry_run=self.dry_run,
            )
            if self.operation_manifest and copied:
                self.operation_manifest.add(
                    action="copy_merge_folder",
                    source_path=original,
                    destination_path=target,
                    status="planned_copy" if self.dry_run else "copied",
                    message="复制合并已有发票文件夹，原文件夹保留",
                    extra={"planned": self.dry_run, "copied_files": copied},
                )

    def _find_invoice_suffix_folder(self, invoice_no: str) -> str:
        if not os.path.isdir(self.output_dir):
            return ""
        display = self.invoice_display(invoice_no)
        separator = self.config.get("folder_name", {}).get("separator", "   ")
        suffix = f"{separator}{display}"
        suffix_pattern = re.compile(re.escape(suffix) + r"(_重复\d+)?$")
        for entry in os.scandir(self.output_dir):
            if entry.is_dir() and suffix_pattern.search(entry.name):
                return entry.path
        return ""

    def _find_prefixed_invoice_folder(self, invoice_no: str) -> str:
        prefix = sanitize_folder_name(invoice_no)
        if not prefix or not os.path.isdir(self.output_dir):
            return ""

        pattern = re.compile(r"^" + re.escape(prefix) + r"(?!\d)", re.IGNORECASE)
        matches = []
        for entry in os.scandir(self.output_dir):
            if entry.is_dir() and pattern.match(entry.name):
                matches.append(entry.path)
        if not matches:
            return ""

        prefix_key = prefix.casefold()

        def sort_key(path: str):
            name = os.path.basename(os.path.normpath(path))
            name_key = name.casefold()
            return (name_key == prefix_key, len(name), name_key)

        return sorted(matches, key=sort_key)[0]

    def invoice_display(self, invoice_no: str) -> str:
        mode = self.config.get("folder_name", {}).get("invoice_display", "digits_only")
        return invoice_display_value(invoice_no, mode)

    def evaluate_invoice(
        self,
        invoice_no: str,
        folder_path: str | None = None,
        allow_prompt: bool = False,
        allow_rename: bool = False,
        save_report: bool = False,
        persist_declaration: bool = True,
        prompt_callback=None,
    ) -> dict:
        with self.lock:
            folder_path = normalize_path(folder_path) if folder_path else ""
            if not folder_path or not os.path.isdir(folder_path):
                folder_path = self.resolve_target_folder(invoice_no)
            if not os.path.isdir(folder_path):
                return {
                    "发票号": invoice_no,
                    "最终发票号显示值": self.invoice_display(invoice_no),
                    "报关号": "",
                    "当前文件夹": "",
                    "最终文件夹": "",
                    "状态": "未找到文件夹",
                    "已识别必需材料": "",
                    "缺少材料": "",
                    "未识别文件": "",
                    "重复材料": "",
                    "需人工确认文件": "",
                    "闭环状态": "",
                    "闭环分数": "",
                    "闭环证据": "",
                    "闭环冲突": "",
                    "闭环未核验": "",
                    "是否已改名": "否",
                    "最后更新时间": time.strftime("%Y-%m-%d %H:%M:%S"),
                    "备注": "输出目录中未找到该发票号文件夹",
                }

            record, classifications = self._build_record(
                invoice_no,
                folder_path,
                persist_declaration=persist_declaration,
            )

            if allow_prompt and record["状态"] == "待输入报关号" and prompt_callback:
                declaration_no = prompt_callback(invoice_no)
                if declaration_no:
                    declaration_no = sanitize_folder_name(declaration_no)
                    if persist_declaration:
                        self.declarations[invoice_no] = declaration_no
                        self._save_declarations()
                    record, classifications = self._build_record(
                        invoice_no,
                        folder_path,
                        persist_declaration=persist_declaration,
                        declaration_override=declaration_no,
                    )
                    record["备注"] = "报关号由用户手动输入"

            if allow_rename and record["状态"] == "可改名":
                record = self._rename_complete_folder(record, classifications)

            if save_report:
                self._save_report_record(record)
            return record

    def process_invoice(
        self,
        invoice_no: str,
        folder_path: str,
        prompt_callback=None,
        allow_rename: bool = True,
    ) -> dict:
        return self.evaluate_invoice(
            invoice_no,
            folder_path,
            allow_prompt=True,
            allow_rename=allow_rename,
            save_report=True,
            persist_declaration=True,
            prompt_callback=prompt_callback,
        )

    def process_all_invoice_folders(
        self,
        invoice_numbers: list[str],
        prompt_callback=None,
        log_callback=None,
        allow_rename: bool = True,
    ):
        for invoice_no in invoice_numbers:
            folder = self.resolve_target_folder(invoice_no)
            if not os.path.isdir(folder):
                continue
            record = self.process_invoice(invoice_no, folder, prompt_callback=prompt_callback, allow_rename=allow_rename)
            if record and log_callback:
                log_callback(f"📋 判断报告: {invoice_no} | {record.get('状态', '')}")

    def _build_record(
        self,
        invoice_no: str,
        folder_path: str,
        persist_declaration: bool = True,
        declaration_override: str = "",
    ) -> tuple[dict, list[dict]]:
        files = [
            os.path.join(folder_path, name)
            for name in os.listdir(folder_path)
            if os.path.isfile(os.path.join(folder_path, name)) and not should_ignore_file(os.path.join(folder_path, name))
        ]
        classifications = [
            self.classifier.classify_file(path, invoice_no)
            for path in files
        ]
        for item in classifications:
            if item["status"] != "matched":
                item["fields"] = {}
                continue
            material = item.get("material") or {}
            if not material.get("file_name_only"):
                if Path(item["path"]).suffix.lower() == ".pdf":
                    item["text"] = self._document_text(item["path"]) or item.get("text", "")
                else:
                    item["text"] = item.get("text") or self._document_text(item["path"])
            item["fields"] = extract_document_fields(item["material_id"], item.get("text", ""), item["filename"])

        required_names = self.classifier.required_names()
        matched_required = {}
        duplicates = []
        unrecognized = []
        manual_review = []
        declaration_text = ""
        purchase_contract_declaration_text = ""

        for item in classifications:
            if item["status"] == "manual_review":
                manual_review.append(item["filename"])
                continue
            if item["status"] != "matched":
                unrecognized.append(item["filename"])
                continue
            if item["material_id"] == "customs_power_of_attorney" and item.get("text"):
                declaration_text += "\n" + item["text"]
            if item["material_id"] == "purchase_contract":
                purchase_contract_declaration_text += "\n" + item["filename"]
            if not item["required"]:
                continue
            name = item["material_name"]
            if name in matched_required:
                duplicates.append(item["filename"])
            else:
                matched_required[name] = item["filename"]

        missing = [name for name in required_names if name not in matched_required]
        declaration_no = (
            declaration_override
            or self.declarations.get(invoice_no, "")
            or extract_declaration_number(declaration_text)
            or extract_declaration_number(purchase_contract_declaration_text)
        )
        if declaration_no:
            declaration_no = sanitize_folder_name(declaration_no)
            if persist_declaration:
                self.declarations[invoice_no] = declaration_no
                self._save_declarations()

        ledger_record = self._ledger_record(invoice_no)
        closed_loop = evaluate_closed_loop(invoice_no, ledger_record, classifications, declaration_no)

        if missing:
            status = "缺材料"
        elif manual_review:
            status = "待复核"
        elif not declaration_no:
            status = "待输入报关号"
        elif closed_loop["status"] in {"strong_pass", "basic_pass"}:
            status = "可改名"
        elif closed_loop["status"] == "review":
            status = "待复核"
        else:
            status = "不通过"

        current_folder_name = os.path.basename(folder_path)
        final_folder = self._final_folder_path(invoice_no, declaration_no) if declaration_no else ""
        if declaration_no and current_folder_name == os.path.basename(final_folder):
            status = "已改名"

        record = {
            "发票号": invoice_no,
            "最终发票号显示值": self.invoice_display(invoice_no),
            "报关号": declaration_no,
            "当前文件夹": folder_path,
            "最终文件夹": final_folder,
            "状态": status,
            "已识别必需材料": "、".join(sorted(matched_required.keys())),
            "缺少材料": "、".join(missing),
            "未识别文件": "、".join(unrecognized),
            "重复材料": "、".join(duplicates),
            "需人工确认文件": "、".join(manual_review),
            "闭环状态": closed_loop["status_label"],
            "闭环分数": closed_loop["score"],
            "闭环证据": "；".join(closed_loop["evidence"]),
            "闭环冲突": "；".join(closed_loop["conflicts"]),
            "闭环未核验": "；".join(closed_loop.get("unverified", [])),
            "是否已改名": "是" if status == "已改名" else "否",
            "最后更新时间": time.strftime("%Y-%m-%d %H:%M:%S"),
            "备注": "",
        }
        return record, classifications

    def _ledger_record(self, invoice_no: str) -> dict:
        if invoice_no in self.ledger_records:
            return self.ledger_records[invoice_no]
        for key, record in self.ledger_records.items():
            if invoice_equivalent(key, invoice_no):
                return record
        return {}

    @staticmethod
    def _document_text(filepath: str) -> str:
        text = pdf_text(filepath, max_pages=10)
        if text:
            return text
        suffix = Path(filepath).suffix.lower()
        if suffix in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".pdf"}:
            return ocr_text(filepath)
        return ""

    def _final_folder_path(self, invoice_no: str, declaration_no: str) -> str:
        separator = self.config.get("folder_name", {}).get("separator", "   ")
        folder_name = sanitize_folder_name(f"{declaration_no}{separator}{self.invoice_display(invoice_no)}")
        return os.path.join(self.output_dir, folder_name)

    def _rename_complete_folder(self, record: dict, classifications: list[dict]) -> dict:
        source_folder = record["当前文件夹"]
        final_folder = self._final_folder_path(record["发票号"], record["报关号"])
        final_folder = unique_path(final_folder) if not paths_equal(source_folder, final_folder) else final_folder
        manifest = {
            "发票号": record["发票号"],
            "报关号": record["报关号"],
            "源文件夹": source_folder,
            "目标文件夹": final_folder,
            "生成时间": time.strftime("%Y-%m-%d %H:%M:%S"),
            "文件": [],
        }

        if self.dry_run or self.copy_first_rename:
            return self._copy_first_complete_folder(record, classifications, source_folder, final_folder)

        try:
            used_targets = set()
            for item in classifications:
                if item["status"] != "matched":
                    continue
                target_name = self._target_filename(item, record["发票号"])
                if not target_name:
                    continue
                target_path = os.path.join(source_folder, target_name)
                if os.path.basename(target_path) in used_targets:
                    target_path = unique_path(target_path)
                elif os.path.exists(target_path) and not paths_equal(item["path"], target_path):
                    target_path = unique_path(target_path)
                used_targets.add(os.path.basename(target_path))
                manifest["文件"].append({
                    "材料类型": item["material_name"],
                    "原文件": item["path"],
                    "目标文件": target_path,
                    "状态": "待改名",
                })

            manifest_path = os.path.join(source_folder, f"rename_manifest_{record['发票号']}.json")
            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(manifest, f, ensure_ascii=False, indent=2)

            for file_record in manifest["文件"]:
                src = file_record["原文件"]
                dst = file_record["目标文件"]
                if paths_equal(src, dst):
                    file_record["状态"] = "无需改名"
                    continue
                os.rename(src, dst)
                file_record["状态"] = "已改名"
                file_record["原文件"] = dst

            with open(manifest_path, "w", encoding="utf-8") as f:
                json.dump(manifest, f, ensure_ascii=False, indent=2)

            if not paths_equal(source_folder, final_folder):
                os.rename(source_folder, final_folder)

            record["当前文件夹"] = final_folder
            record["最终文件夹"] = final_folder
            record["状态"] = "已改名"
            record["是否已改名"] = "是"
            record["备注"] = "已按报关号和发票号完成文件夹与文件改名"
        except Exception as exc:
            record["状态"] = "改名失败"
            record["备注"] = str(exc)
            logging.error(f"退税材料改名失败: {record['发票号']} | {exc}")
        return record

    def _copy_first_complete_folder(
        self,
        record: dict,
        classifications: list[dict],
        source_folder: str,
        final_folder: str,
    ) -> dict:
        try:
            file_records = self._copy_first_rename_plan(record, classifications, source_folder, final_folder)
            if self.dry_run:
                for file_record in file_records:
                    record_file_copy(
                        self.operation_manifest,
                        "rename_copy_file",
                        file_record["source"],
                        file_record["destination"],
                        "planned_copy",
                        invoice_no=record["发票号"],
                        subfolder=os.path.basename(os.path.normpath(final_folder)),
                        message="预览生成标准命名副本",
                        planned=True,
                    )
                if self.operation_manifest:
                    self.operation_manifest.add(
                        action="rename_copy_folder",
                        source_path=source_folder,
                        destination_path=final_folder,
                        status="planned_copy",
                        invoice_no=record["发票号"],
                        message="预览生成标准命名副本，原文件夹保留",
                        extra={"planned": True, "file_count": len(file_records)},
                    )
                record["最终文件夹"] = final_folder
                record["状态"] = "预览改名"
                record["是否已改名"] = "否"
                record["备注"] = "预览模式：未复制、未改名、未移动原文件夹"
                return record

            for file_record in file_records:
                copy_file_with_manifest(
                    file_record["source"],
                    file_record["destination"],
                    operation_manifest=self.operation_manifest,
                    action="rename_copy_file",
                    invoice_no=record["发票号"],
                    subfolder=os.path.basename(os.path.normpath(final_folder)),
                    message="生成标准命名副本",
                )
            if self.operation_manifest:
                self.operation_manifest.add(
                    action="rename_copy_folder",
                    source_path=source_folder,
                    destination_path=final_folder,
                    status="copied",
                    invoice_no=record["发票号"],
                    message="已生成标准命名副本，原文件夹保留",
                    extra={"planned": False, "file_count": len(file_records)},
                )

            record["当前文件夹"] = source_folder
            record["最终文件夹"] = final_folder
            record["状态"] = "已生成改名副本"
            record["是否已改名"] = "否"
            record["备注"] = "已按报关号和发票号生成标准命名副本，原文件夹未移动"
        except Exception as exc:
            record["状态"] = "改名失败"
            record["备注"] = str(exc)
            logging.error(f"退税材料生成改名副本失败: {record['发票号']} | {exc}")
        return record

    def _copy_first_rename_plan(
        self,
        record: dict,
        classifications: list[dict],
        source_folder: str,
        final_folder: str,
    ) -> list[dict]:
        classification_by_path = {
            comparable_path(item["path"]): item
            for item in classifications
            if item.get("path")
        }
        used_targets = set()
        file_records = []

        for name in sorted(os.listdir(source_folder)):
            source_path = os.path.join(source_folder, name)
            if not os.path.isfile(source_path) or should_ignore_file(source_path):
                continue
            item = classification_by_path.get(comparable_path(source_path))
            target_name = name
            material_name = ""
            if item and item.get("status") == "matched":
                target_name = self._target_filename(item, record["发票号"]) or name
                material_name = item.get("material_name", "")
            destination = self._reserve_copy_destination(
                source_path,
                os.path.join(final_folder, target_name),
                used_targets,
            )
            file_records.append({
                "source": source_path,
                "destination": destination,
                "material_name": material_name,
            })
        return file_records

    def _reserve_copy_destination(self, source_path: str, destination: str, used_targets: set) -> str:
        folder = os.path.dirname(destination)
        stem = Path(destination).stem
        suffix = Path(destination).suffix
        counter = 0
        while True:
            candidate_base = destination if counter == 0 else os.path.join(folder, f"{stem}_重复{counter}{suffix}")
            candidate, _ = resolve_copy_destination(source_path, candidate_base)
            key = comparable_path(candidate)
            if key not in used_targets:
                used_targets.add(key)
                return candidate
            counter += 1

    def _target_filename(self, item: dict, invoice_no: str) -> str:
        suffix = Path(item["filename"]).suffix
        material = item["material"] or {}
        if item["material_id"] == "purchase_contract":
            return strip_invoice_prefix(item["filename"], invoice_no)
        return f"{material.get('final_name') or item['material_name']}{suffix}"

    def _load_declarations(self) -> dict:
        if not os.path.exists(self.declaration_path):
            return {}
        try:
            with open(self.declaration_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_declarations(self):
        tmp_path = f"{self.declaration_path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(self.declarations, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, self.declaration_path)

    def _save_report_record(self, record: dict):
        today = date.today().strftime("%Y-%m-%d")
        json_path = os.path.join(self.output_dir, f"判断报告_{today}.json")
        records = []
        if os.path.exists(json_path):
            try:
                with open(json_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    records = data
            except json.JSONDecodeError:
                backup = f"{json_path}.bak_{int(time.time())}"
                os.replace(json_path, backup)

        by_invoice = {
            rec.get("发票号"): rec
            for rec in records
            if isinstance(rec, dict) and rec.get("发票号")
        }
        by_invoice[record["发票号"]] = record
        records = list(by_invoice.values())

        tmp_path = f"{json_path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, json_path)
        self._save_report_excel(records, os.path.join(self.output_dir, f"判断报告_{today}.xlsx"))

    def _save_report_excel(self, records: list[dict], path: str):
        if openpyxl is None:
            return
        headers = [
            "发票号", "最终发票号显示值", "报关号", "当前文件夹", "最终文件夹",
            "状态", "已识别必需材料", "缺少材料", "未识别文件", "重复材料",
            "需人工确认文件", "闭环状态", "闭环分数", "闭环证据", "闭环冲突", "闭环未核验",
            "是否已改名", "最后更新时间", "备注",
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "判断报告"
        ws.append(headers)
        ExcelLogger._style_header(ws, headers)
        for record in sorted(records, key=lambda item: item.get("发票号", "")):
            ws.append([record.get(header, "") for header in headers])
        widths = [16, 18, 24, 45, 45, 14, 28, 28, 38, 28, 38, 16, 12, 55, 45, 45, 12, 22, 40]
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width
        wb.save(path)
        wb.close()


# ─────────────────────────────────────────────
# 工作日志管理器
# ─────────────────────────────────────────────

class ExcelLogger:
    def __init__(self, output_dir: str):
        self.output_dir = output_dir
        self.lock = threading.Lock()
        self._records: list[dict] = []

    def add_record(
        self,
        filename: str,
        subfolder: str,
        status: str,
        invoice_no: str = "",
        source_path: str = "",
        dest_path: str = "",
        message: str = "",
    ):
        with self.lock:
            self._records.append({
                "记录ID": f"{time.time_ns()}-{threading.get_ident()}",
                "文件名": filename,
                "归入子文件夹": subfolder,
                "处理日期": date.today().strftime("%Y-%m-%d"),
                "处理时间": time.strftime("%Y-%m-%d %H:%M:%S"),
                "发票号": invoice_no,
                "状态": status,
                "源路径": source_path,
                "目标路径": dest_path,
                "说明": message,
            })

    def save(self):
        """将新增记录写入 JSON 工作日志，并同步追加到 Excel 日志。"""
        with self.lock:
            records = list(self._records)
            if not records:
                return

            try:
                self._append_json_records(records)
                self._append_excel_records(records)
            except Exception as e:
                logging.error(f"工作日志保存失败: {e}")
                return

            del self._records[:len(records)]

    def _append_json_records(self, records: list[dict]):
        today_str = date.today().strftime("%Y-%m-%d")
        log_path = os.path.join(self.output_dir, f"工作日志_{today_str}.json")
        existing_records = []

        if os.path.exists(log_path):
            try:
                with open(log_path, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list):
                    existing_records = data
            except json.JSONDecodeError:
                backup_path = f"{log_path}.bak_{int(time.time())}"
                os.replace(log_path, backup_path)
                logging.error(f"JSON 工作日志损坏，已备份为: {backup_path}")

        seen_ids = {
            rec.get("记录ID")
            for rec in existing_records
            if isinstance(rec, dict) and rec.get("记录ID")
        }
        for rec in records:
            if rec["记录ID"] not in seen_ids:
                existing_records.append(rec)
                seen_ids.add(rec["记录ID"])

        tmp_path = f"{log_path}.tmp"
        with open(tmp_path, "w", encoding="utf-8") as f:
            json.dump(existing_records, f, ensure_ascii=False, indent=2)
        os.replace(tmp_path, log_path)

    def _append_excel_records(self, records: list[dict]):
        if openpyxl is None:
            return

        today_str = date.today().strftime("%Y-%m-%d")
        log_path = os.path.join(self.output_dir, f"工作日志_{today_str}.xlsx")

        if os.path.exists(log_path):
            wb = openpyxl.load_workbook(log_path)
            ws_detail = wb["处理明细"] if "处理明细" in wb.sheetnames else wb.create_sheet("处理明细")
            ws_summary = wb["汇总统计"] if "汇总统计" in wb.sheetnames else wb.create_sheet("汇总统计")
        else:
            wb = openpyxl.Workbook()
            ws_detail = wb.active
            ws_detail.title = "处理明细"
            ws_summary = wb.create_sheet("汇总统计")

        # ── 处理明细 Sheet ──
        headers = ["序号", "文件名", "归入子文件夹", "处理时间", "状态", "源路径", "目标路径", "说明"]
        first_row_has_values = any(cell.value for cell in ws_detail[1])
        if ws_detail.max_row <= 1 and not first_row_has_values:
            for col_idx, header in enumerate(headers, start=1):
                ws_detail.cell(row=1, column=col_idx, value=header)
            self._style_header(ws_detail, headers)
        else:
            for col_idx, header in enumerate(headers, start=1):
                if not ws_detail.cell(row=1, column=col_idx).value:
                    ws_detail.cell(row=1, column=col_idx, value=header)
            self._style_header(ws_detail, headers)

        start_row = ws_detail.max_row + 1
        for i, rec in enumerate(records, start=start_row):
            ws_detail.append([
                i - 1,
                rec["文件名"],
                rec["归入子文件夹"],
                rec["处理时间"],
                rec["状态"],
                rec.get("源路径", ""),
                rec.get("目标路径", ""),
                rec.get("说明", ""),
            ])

        # 设置列宽
        col_widths = [8, 45, 20, 22, 40, 65, 65, 70]
        for col_idx, width in enumerate(col_widths, start=1):
            ws_detail.column_dimensions[
                openpyxl.utils.get_column_letter(col_idx)
            ].width = width

        # ── 汇总统计 Sheet ──
        all_records = []
        for row in ws_detail.iter_rows(min_row=2, values_only=True):
            if row[2]:
                all_records.append(row[2])

        folder_count: dict = {}
        for sf in all_records:
            folder_count[sf] = folder_count.get(sf, 0) + 1

        ws_summary.delete_rows(1, ws_summary.max_row + 1)
        summary_headers = ["子文件夹名称", "文件数量"]
        ws_summary.append(summary_headers)
        self._style_header(ws_summary, summary_headers)

        for sf, cnt in sorted(folder_count.items(), key=lambda x: -x[1]):
            ws_summary.append([sf, cnt])

        total = sum(folder_count.values())
        total_row = ws_summary.max_row + 1
        ws_summary.cell(row=total_row, column=1, value="【合计】")
        ws_summary.cell(row=total_row, column=2, value=total)
        bold = Font(bold=True, color="C0392B")
        ws_summary.cell(row=total_row, column=1).font = bold
        ws_summary.cell(row=total_row, column=2).font = bold

        ws_summary.column_dimensions["A"].width = 22
        ws_summary.column_dimensions["B"].width = 14

        wb.save(log_path)
        wb.close()

    @staticmethod
    def _style_header(ws, headers):
        header_fill = PatternFill("solid", fgColor="2E86AB")
        header_font = Font(bold=True, color="FFFFFF")
        header_align = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = border


def backup_excel_file(path: str) -> str:
    source = Path(path)
    if not source.exists():
        raise FileNotFoundError(f"Excel 文件不存在：{source}")
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    backup_path = f"{source}.bak_{timestamp}"
    shutil.copy2(source, backup_path)
    return backup_path


def ensure_columns(ws, headers: list[str]) -> dict:
    existing = {}
    max_col = max(ws.max_column, 1)
    for col_idx in range(1, max_col + 1):
        header = format_excel_cell(ws.cell(row=1, column=col_idx).value)
        if header:
            existing[header] = col_idx

    next_col = max_col + 1
    for header in headers:
        if header not in existing:
            ws.cell(row=1, column=next_col, value=header)
            existing[header] = next_col
            next_col += 1
    return {header: existing[header] for header in headers if header in existing}


def _invoice_header_index(headers: list[str], invoice_column_name: str = "发票号") -> int | None:
    aliases = {invoice_column_name, "发票号", "发票号码", "Invoice", "invoice", "INV"}
    for idx, header in enumerate(headers, start=1):
        if header in aliases:
            return idx
    return None


def _status_record_for_invoice(invoice_no: str, status_by_invoice: dict[str, dict]) -> dict | None:
    normalized = normalize_identifier(invoice_no)
    digits = invoice_digits(normalized)
    for key, record in status_by_invoice.items():
        key_norm = normalize_identifier(key)
        if normalized and normalized == key_norm:
            return record
        key_digits = invoice_digits(key_norm)
        if digits and len(digits) >= 6 and digits == key_digits:
            return record
    for key, record in status_by_invoice.items():
        if invoice_equivalent(invoice_no, key):
            return record
    return None


def _status_cell_values(record: dict) -> dict:
    return {
        "退税齐套状态": record.get("退税齐套状态") or record.get("tax_status") or record.get("状态", ""),
        "财务批次": record.get("财务批次") or record.get("batch_name", ""),
        "齐套检查时间": record.get("齐套检查时间") or record.get("最后更新时间", ""),
        "缺少材料": record.get("缺少材料", ""),
        "闭环状态": record.get("闭环状态", ""),
        "闭环分数": record.get("闭环分数", ""),
        "退税包路径": record.get("退税包路径") or record.get("批次包文件夹", ""),
        "备注": record.get("备注", ""),
    }


def write_status_to_workbook(
    workbook_path: str,
    status_by_invoice: dict[str, dict],
    invoice_column_name: str = "发票号",
    create_backup: bool = True,
) -> str:
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法回写 Excel")
    if not status_by_invoice:
        return ""

    workbook_path = normalize_path(workbook_path)
    with EXCEL_WRITE_LOCK:
        backup_path = backup_excel_file(workbook_path) if create_backup else ""
        wb = None
        tmp_path = f"{workbook_path}.tmp_{time.strftime('%Y%m%d_%H%M%S')}_{threading.get_ident()}.xlsx"
        status_headers = [
            "退税齐套状态",
            "财务批次",
            "齐套检查时间",
            "缺少材料",
            "闭环状态",
            "闭环分数",
            "退税包路径",
            "备注",
        ]

        try:
            wb = openpyxl.load_workbook(workbook_path)
            for ws in wb.worksheets:
                headers = [
                    format_excel_cell(ws.cell(row=1, column=col_idx).value)
                    for col_idx in range(1, max(ws.max_column, 1) + 1)
                ]
                invoice_col = _invoice_header_index(headers, invoice_column_name)
                data_start_row = 2
                if invoice_col is None:
                    ws.insert_rows(1)
                    ws.cell(row=1, column=1, value=invoice_column_name)
                    invoice_col = 1
                    data_start_row = 2

                columns = ensure_columns(ws, status_headers)
                for row_idx in range(data_start_row, ws.max_row + 1):
                    invoice_no = format_excel_cell(ws.cell(row=row_idx, column=invoice_col).value)
                    if not invoice_no:
                        continue
                    status_record = _status_record_for_invoice(invoice_no, status_by_invoice)
                    if not status_record:
                        continue
                    values = _status_cell_values(status_record)
                    for header in status_headers:
                        ws.cell(row=row_idx, column=columns[header], value=values.get(header, ""))

            wb.save(tmp_path)
            wb.close()
            wb = None
            os.replace(tmp_path, workbook_path)
            return backup_path
        except (PermissionError, OSError) as exc:
            raise RuntimeError("Excel 文件可能正在被打开，请关闭后重试") from exc
        finally:
            if wb is not None:
                wb.close()
            if os.path.exists(tmp_path):
                try:
                    os.remove(tmp_path)
                except OSError:
                    pass


def load_finance_batch_invoices(batch_workbook_path: str) -> list[dict]:
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法读取财务批次表")

    records = []
    seen = set()
    wb = openpyxl.load_workbook(batch_workbook_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [format_excel_cell(value) for value in rows[0]]
            invoice_idx = None
            for idx, header in enumerate(headers):
                if header == "发票号":
                    invoice_idx = idx
                    break
            start_idx = 1 if invoice_idx is not None else 0
            if invoice_idx is None:
                invoice_idx = 0

            for offset, row in enumerate(rows[start_idx:], start=start_idx + 1):
                invoice_no = format_excel_cell(row[invoice_idx] if invoice_idx < len(row) else "").strip()
                if not invoice_no:
                    continue
                key = invoice_dedup_key(invoice_no)
                if key in seen:
                    continue
                seen.add(key)
                records.append({
                    "invoice_no": invoice_no,
                    "sheet": ws.title,
                    "row": offset,
                })
    finally:
        wb.close()
    return records


def tax_status_from_record(record: dict, in_ledger: bool, folder_exists: bool) -> str:
    if not in_ledger:
        return "不在主台账"
    if not folder_exists:
        return "未找到文件夹"
    if str(record.get("缺少材料", "")).strip():
        return "缺材料"
    if str(record.get("需人工确认文件", "")).strip():
        return "待复核"

    status = str(record.get("状态", "")).strip()
    if status == "不通过":
        return "不通过"
    if status in {"待输入报关号", "待复核", "改名失败"}:
        return "待复核"
    if record.get("闭环状态") in {"强闭环通过", "基本闭环通过"}:
        return "齐了"
    return "待复核"


def _batch_summary_counts(records: list[dict]) -> dict:
    keys = ["齐了", "缺材料", "待复核", "未找到文件夹", "不在主台账", "不通过"]
    counts = {key: 0 for key in keys}
    other = 0
    for record in records:
        status = record.get("退税齐套状态", "")
        if status in counts:
            counts[status] += 1
        else:
            other += 1
    counts["其他"] = other
    return counts


def save_batch_report_json(records: list[dict], path: str):
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def save_batch_report_excel(records: list[dict], path: str):
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法生成批次统计报告")

    headers = [
        "序号",
        "发票号",
        "是否在主台账",
        "是否找到文件夹",
        "退税齐套状态",
        "原文件夹",
        "批次包文件夹",
        "已识别必需材料",
        "缺少材料",
        "未识别文件",
        "重复材料",
        "需人工确认文件",
        "闭环状态",
        "闭环分数",
        "闭环证据",
        "闭环冲突",
        "闭环未核验",
        "报关号",
        "最后更新时间",
        "备注",
    ]
    Path(path).parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws_detail = wb.active
    ws_detail.title = "批次明细"
    ws_detail.append(headers)
    ExcelLogger._style_header(ws_detail, headers)
    for record in records:
        ws_detail.append([record.get(header, "") for header in headers])

    widths = [8, 18, 14, 14, 16, 45, 45, 28, 28, 38, 28, 38, 16, 12, 55, 45, 24, 22, 45]
    for col_idx, width in enumerate(widths, start=1):
        ws_detail.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws_summary = wb.create_sheet("汇总统计")
    summary_headers = ["项目", "数量"]
    ws_summary.append(summary_headers)
    ExcelLogger._style_header(ws_summary, summary_headers)
    counts = _batch_summary_counts(records)
    summary_rows = [
        ("总发票数", len(records)),
        ("齐了", counts["齐了"]),
        ("缺材料", counts["缺材料"]),
        ("待复核", counts["待复核"]),
        ("未找到文件夹", counts["未找到文件夹"]),
        ("不在主台账", counts["不在主台账"]),
        ("不通过", counts["不通过"]),
        ("其他", counts["其他"]),
    ]
    for row in summary_rows:
        ws_summary.append(row)
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 12

    tmp_path = f"{path}.tmp"
    wb.save(tmp_path)
    wb.close()
    os.replace(tmp_path, path)


def _find_equivalent_invoice_key(records: dict[str, dict], invoice_no: str) -> str:
    if invoice_no in records:
        return invoice_no
    for key in records:
        if invoice_equivalent(key, invoice_no):
            return key
    return ""


def copy_invoice_folder(
    source_folder: str,
    batch_output_dir: str,
    invoice_no: str,
    record: dict,
    material_classifier=None,
) -> str:
    del record
    source_folder = normalize_path(source_folder)
    batch_output_dir = normalize_path(batch_output_dir)
    os.makedirs(batch_output_dir, exist_ok=True)
    target_folder = unique_path(os.path.join(batch_output_dir, os.path.basename(source_folder)))

    def ignore_batch_irrelevant(directory, names):
        ignored = []
        for name in names:
            path = os.path.join(directory, name)
            if (
                name == "declaration_overrides.json"
                or name.startswith(("工作日志_", "判断报告_", "rename_manifest_"))
                or should_ignore_file(path)
            ):
                ignored.append(name)
                continue
            if material_classifier and os.path.isfile(path):
                try:
                    classification = material_classifier.classify_file(path, invoice_no)
                except Exception:
                    ignored.append(name)
                    continue
                if classification.get("status") not in ROUTABLE_MATERIAL_STATUSES:
                    ignored.append(name)
        return ignored

    shutil.copytree(source_folder, target_folder, ignore=ignore_batch_irrelevant)
    return target_folder


class FinanceBatchProcessor:
    def __init__(
        self,
        output_dir: str,
        company_ledger_path: str,
        tax_processor: TaxRefundProcessor,
        settings: dict,
        log_callback=None,
    ):
        self.output_dir = normalize_path(output_dir)
        self.company_ledger_path = normalize_path(company_ledger_path)
        self.tax_processor = tax_processor
        self.settings = settings or default_app_settings()
        self.log_callback = log_callback

    def _log(self, message: str):
        if self.log_callback:
            self.log_callback(message)

    def build_batch_output_dir(self, batch_workbook_path: str) -> str:
        finance_settings = self.settings.get("finance_batch", {})
        output_subdir = (
            finance_settings.get("output_subdir")
            or finance_settings.get("default_output_subdir")
            or "财务退税批次包"
        )
        batch_name = sanitize_folder_name(Path(batch_workbook_path).stem)
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        return os.path.join(self.output_dir, output_subdir, f"{batch_name}_{timestamp}")

    def process_batch(self, batch_workbook_path: str) -> dict:
        batch_workbook_path = normalize_path(batch_workbook_path)
        if not os.path.isfile(batch_workbook_path):
            raise FileNotFoundError(f"财务批次表不存在：{batch_workbook_path}")

        batch_items = load_finance_batch_invoices(batch_workbook_path)
        if not batch_items:
            raise RuntimeError("财务批次表中没有读取到发票号")

        ledger_records = load_invoice_records(self.company_ledger_path)
        self.tax_processor.ledger_records = ledger_records
        batch_output_dir = self.build_batch_output_dir(batch_workbook_path)
        os.makedirs(batch_output_dir, exist_ok=True)
        batch_name = os.path.basename(batch_output_dir)

        self._log(f"📦 财务批次处理开始：{Path(batch_workbook_path).name}，共 {len(batch_items)} 个发票号")
        records = []
        status_by_invoice = {}

        for index, item in enumerate(batch_items, start=1):
            requested_invoice = item["invoice_no"].strip()
            ledger_key = _find_equivalent_invoice_key(ledger_records, requested_invoice)
            in_ledger = bool(ledger_key)
            eval_invoice = ledger_key or requested_invoice
            source_folder = self.tax_processor.resolve_target_folder(eval_invoice)
            folder_exists = os.path.isdir(source_folder)

            record = self.tax_processor.evaluate_invoice(
                eval_invoice,
                source_folder if folder_exists else None,
                allow_prompt=False,
                allow_rename=False,
                save_report=False,
                persist_declaration=False,
            )
            copied_folder = ""
            if folder_exists:
                copied_folder = copy_invoice_folder(
                    source_folder,
                    batch_output_dir,
                    eval_invoice,
                    record,
                    material_classifier=self.tax_processor.classifier,
                )

            tax_status = tax_status_from_record(record, in_ledger=in_ledger, folder_exists=folder_exists)
            remarks = [record.get("备注", "")]
            if not in_ledger:
                remarks.append("财务批次发票号不在公司系统主台账")
            remarks = "；".join(item for item in remarks if item)

            batch_record = {
                "序号": index,
                "发票号": requested_invoice,
                "是否在主台账": "是" if in_ledger else "否",
                "是否找到文件夹": "是" if folder_exists else "否",
                "退税齐套状态": tax_status,
                "原文件夹": source_folder if folder_exists else "",
                "批次包文件夹": copied_folder,
                "退税包路径": copied_folder,
                "财务批次": batch_name,
                "齐套检查时间": record.get("最后更新时间", ""),
                "已识别必需材料": record.get("已识别必需材料", ""),
                "缺少材料": record.get("缺少材料", ""),
                "未识别文件": record.get("未识别文件", ""),
                "重复材料": record.get("重复材料", ""),
                "需人工确认文件": record.get("需人工确认文件", ""),
                "闭环状态": record.get("闭环状态", ""),
                "闭环分数": record.get("闭环分数", ""),
                "闭环证据": record.get("闭环证据", ""),
                "闭环冲突": record.get("闭环冲突", ""),
                "闭环未核验": record.get("闭环未核验", ""),
                "报关号": record.get("报关号", ""),
                "最后更新时间": record.get("最后更新时间", ""),
                "备注": remarks,
                "batch_source_sheet": item.get("sheet", ""),
                "batch_source_row": item.get("row", ""),
            }
            records.append(batch_record)
            status_by_invoice[requested_invoice] = batch_record
            status_by_invoice[eval_invoice] = batch_record

            if index % 20 == 0 or index == len(batch_items):
                self._log(f"📦 财务批次进度：{index}/{len(batch_items)}")

        report_xlsx = os.path.join(batch_output_dir, "批次统计报告.xlsx")
        report_json = os.path.join(batch_output_dir, "批次统计报告.json")
        save_batch_report_excel(records, report_xlsx)
        save_batch_report_json(records, report_json)

        finance_settings = self.settings.get("finance_batch", {})
        marked_batch_copy = ""
        if finance_settings.get("create_marked_batch_copy", True):
            marked_batch_copy = os.path.join(
                batch_output_dir,
                f"{sanitize_folder_name(Path(batch_workbook_path).stem)}_已标注.xlsx",
            )
            shutil.copy2(batch_workbook_path, marked_batch_copy)
            write_status_to_workbook(marked_batch_copy, status_by_invoice, create_backup=False)

        company_backup = ""
        finance_backup = ""
        if finance_settings.get("write_back_company_ledger", True):
            company_backup = write_status_to_workbook(self.company_ledger_path, status_by_invoice)
        if finance_settings.get("write_back_finance_batch", True):
            finance_backup = write_status_to_workbook(batch_workbook_path, status_by_invoice)

        finance_settings["last_batch_workbook_path"] = batch_workbook_path
        finance_settings["last_batch_name"] = batch_name
        finance_settings["last_batch_output_dir"] = batch_output_dir
        self.settings["finance_batch"] = finance_settings
        save_app_settings(self.settings, APP_SETTINGS_PATH)

        counts = _batch_summary_counts(records)
        summary = {
            "batch_name": batch_name,
            "total": len(records),
            "counts": counts,
            "batch_output_dir": batch_output_dir,
            "report_xlsx": report_xlsx,
            "report_json": report_json,
            "marked_batch_copy": marked_batch_copy,
            "company_ledger_backup": company_backup,
            "finance_batch_backup": finance_backup,
        }
        self._log(f"📦 财务批次处理完成：{batch_output_dir}")
        return summary


def sync_single_invoice_status_to_company_ledger(
    invoice_no: str,
    record: dict,
    company_ledger_path: str,
    batch_name: str = "",
    create_backup: bool = True,
):
    ledger_records = load_invoice_records(company_ledger_path)
    in_ledger = bool(_find_equivalent_invoice_key(ledger_records, invoice_no))
    folder_exists = bool(record.get("当前文件夹") and os.path.isdir(record.get("当前文件夹")))
    tax_status = tax_status_from_record(record, in_ledger=in_ledger, folder_exists=folder_exists)
    status_record = dict(record)
    status_record["退税齐套状态"] = tax_status
    status_record["财务批次"] = batch_name
    status_record["齐套检查时间"] = record.get("最后更新时间", "")
    return write_status_to_workbook(company_ledger_path, {invoice_no: status_record}, create_backup=create_backup)


class CompanyLedgerSyncManager:
    def __init__(
        self,
        company_ledger_path: str,
        batch_name: str = "",
        log_callback=None,
        debounce_seconds: float = 5.0,
    ):
        self.company_ledger_path = company_ledger_path
        self.batch_name = batch_name
        self.log_callback = log_callback
        self.debounce_seconds = max(0.1, float(debounce_seconds))
        self._pending: dict[str, dict] = {}
        self._timer: threading.Timer | None = None
        self._backup_dates: set[str] = set()
        self._lock = threading.Lock()

    def enqueue(self, invoice_no: str, record: dict):
        if not invoice_no or not record:
            return
        with self._lock:
            self._pending[invoice_no] = dict(record)
            if self._timer:
                self._timer.cancel()
            self._timer = threading.Timer(self.debounce_seconds, self.flush)
            self._timer.daemon = True
            self._timer.start()

    def flush(self):
        with self._lock:
            timer = self._timer
            self._timer = None
            if timer:
                timer.cancel()
            pending = dict(self._pending)
            self._pending.clear()

        if not pending:
            return

        try:
            ledger_records = load_invoice_records(self.company_ledger_path)
            status_by_invoice = {}
            for invoice_no, record in pending.items():
                in_ledger = bool(_find_equivalent_invoice_key(ledger_records, invoice_no))
                folder_exists = bool(record.get("当前文件夹") and os.path.isdir(record.get("当前文件夹")))
                status_record = dict(record)
                status_record["退税齐套状态"] = tax_status_from_record(
                    record,
                    in_ledger=in_ledger,
                    folder_exists=folder_exists,
                )
                status_record["财务批次"] = self.batch_name
                status_record["齐套检查时间"] = record.get("最后更新时间", "")
                status_by_invoice[invoice_no] = status_record

            today = date.today().isoformat()
            create_backup = today not in self._backup_dates
            backup_path = write_status_to_workbook(
                self.company_ledger_path,
                status_by_invoice,
                create_backup=create_backup,
            )
            if create_backup:
                self._backup_dates.add(today)
            if self.log_callback:
                backup_text = f"，备份: {backup_path}" if backup_path else ""
                self.log_callback(f"📒 公司主台账状态已批量同步 {len(status_by_invoice)} 票{backup_text}")
        except Exception as exc:
            if self.log_callback:
                self.log_callback(f"⚠ 公司主台账状态同步失败: {exc}")

    def close(self, flush: bool = True):
        if flush:
            self.flush()
            return
        with self._lock:
            if self._timer:
                self._timer.cancel()
                self._timer = None
            self._pending.clear()


# ─────────────────────────────────────────────
# 文件监控处理器
# ─────────────────────────────────────────────

class FileHandler(FileSystemEventHandler if Observer else object):
    def __init__(self, output_dir: str, excel_logger: ExcelLogger,
                 invoice_matcher: InvoiceMatcher, executor: ThreadPoolExecutor,
                 log_callback, pending_set: set, tax_processor: TaxRefundProcessor = None,
                 declaration_prompt_callback=None, company_ledger_path: str = "",
                 finance_batch_name: str = "", ledger_sync_manager: CompanyLedgerSyncManager = None,
                 watch_dir: str = "", stop_event: threading.Event = None,
                 dry_run: bool = False, operation_manifest: OperationManifest | None = None,
                 auto_rename: bool = False):
        if Observer:
            super().__init__()
        self.output_dir = output_dir
        self.excel_logger = excel_logger
        self.invoice_matcher = invoice_matcher
        self.executor = executor
        self.log_callback = log_callback
        self.pending_set = pending_set   # 正在处理的文件路径集合
        self.tax_processor = tax_processor
        self.declaration_prompt_callback = declaration_prompt_callback
        self.company_ledger_path = company_ledger_path
        self.finance_batch_name = finance_batch_name
        self.ledger_sync_manager = ledger_sync_manager
        self.watch_dir = normalize_path(watch_dir) if watch_dir else ""
        self.stop_event = stop_event
        self.dry_run = bool(dry_run)
        self.operation_manifest = operation_manifest
        self.auto_rename = bool(auto_rename)
        self._lock = threading.Lock()

    def on_closed(self, event):
        """文件关闭（写入完成）时触发（watchdog >= 2.1）"""
        if not event.is_directory:
            self._schedule(event.src_path)

    def on_created(self, event):
        if not event.is_directory:
            self._schedule(event.src_path)

    def on_moved(self, event):
        if not event.is_directory:
            self._schedule(event.dest_path)

    def _schedule(self, filepath: str):
        if self.stop_event and self.stop_event.is_set():
            return False
        if should_ignore_file(filepath):
            return False
        filepath = normalize_path(filepath)
        if is_path_inside(filepath, self.output_dir):
            return False
        if self.watch_dir and not self.invoice_matcher.match_path(filepath, source_root=self.watch_dir):
            return False
        with self._lock:
            if filepath in self.pending_set:
                return False
            self.pending_set.add(filepath)
        try:
            self.executor.submit(self._process, filepath)
            return True
        except RuntimeError:
            with self._lock:
                self.pending_set.discard(filepath)
            return False

    def _process(self, filepath: str):
        filename = os.path.basename(filepath)
        try:
            if self.stop_event and self.stop_event.is_set():
                return
            # 等待文件写入完成
            ready, reason = wait_for_file_ready(
                filepath,
                timeout_seconds=30.0,
                poll_interval=1.0,
                stable_checks=2,
                stop_event=self.stop_event,
            )
            if self.stop_event and self.stop_event.is_set():
                return
            if not ready:
                msg = f"文件未就绪，已放弃: {filename} | {reason}"
                self.log_callback(f"⚠ {msg}")
                self.excel_logger.add_record(
                    filename,
                    "",
                    "失败",
                    source_path=filepath,
                    message=msg,
                )
                self.excel_logger.save()
                return

            success, subfolder, msg, dest_path, invoice_no, status = move_file_to_output(
                filepath,
                self.output_dir,
                self.invoice_matcher,
                invoice_dir_resolver=(
                    getattr(self.tax_processor, "resolve_target_folder", None)
                    if self.tax_processor
                    else None
                ),
                invoice_dir_planner=(
                    getattr(self.tax_processor, "plan_target_folder", None)
                    if self.tax_processor
                    else None
                ),
                source_root=self.watch_dir,
                material_classifier=getattr(self.tax_processor, "classifier", None),
                unmatched_action="skip",
                stop_event=self.stop_event,
                dry_run=self.dry_run,
                operation_manifest=self.operation_manifest,
            )
            if success and status == "成功":
                self.log_callback(f"✅ {msg}", processed=True)
                self.log_callback(f"   复制路径: {filepath} -> {dest_path}")
            elif success and status == STATUS_PLANNED:
                self.log_callback(f"🧾 {msg}", processed=True)
                self.log_callback(f"   计划路径: {filepath} -> {dest_path}")
            elif success and status == STATUS_SKIPPED:
                self.log_callback(f"⏭ {msg}")
                self.log_callback(f"   保留位置: {filepath}")
            elif success:
                self.log_callback(f"⚠ {msg}", processed=True)
                if dest_path:
                    self.log_callback(f"   复制路径: {filepath} -> {dest_path}")
            else:
                self.log_callback(f"❌ {msg}")
            self.excel_logger.add_record(
                filename,
                subfolder,
                status if success else msg,
                invoice_no=invoice_no,
                source_path=filepath,
                dest_path=dest_path,
                message=msg,
            )
            # 每处理一个文件就保存一次日志（保证实时性）
            self.excel_logger.save()

            if (
                success
                and status == "成功"
                and invoice_no
                and self.tax_processor
                and not (self.stop_event and self.stop_event.is_set())
            ):
                record = self.tax_processor.process_invoice(
                    invoice_no,
                    os.path.dirname(dest_path),
                    prompt_callback=self.declaration_prompt_callback,
                    allow_rename=self.auto_rename,
                )
                if record:
                    self.log_callback(
                        f"📋 退税材料判断: {invoice_no} | {record.get('状态', '')} | "
                        f"缺少: {record.get('缺少材料') or '无'}"
                    )
                    if self.ledger_sync_manager:
                        try:
                            self.ledger_sync_manager.enqueue(invoice_no, record)
                        except Exception as sync_exc:
                            self.log_callback(f"⚠ 公司主台账状态同步失败: {invoice_no} | {sync_exc}")
        except Exception as e:
            msg = f"异常: {filepath} | {e}"
            self.log_callback(f"❌ {msg}")
            self.excel_logger.add_record(
                filename,
                "",
                "失败",
                source_path=filepath,
                message=msg,
            )
            self.excel_logger.save()
        finally:
            with self._lock:
                self.pending_set.discard(filepath)


# ─────────────────────────────────────────────
# 图形界面
# ─────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("📁 文件自动归类助手")
        self.geometry("780x600")
        self.resizable(True, True)
        self.configure(bg="#F0F4F8")

        # 状态变量
        self.watch_dir = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.dry_run_var = tk.BooleanVar(value=True)
        self.auto_rename_var = tk.BooleanVar(value=False)
        self.is_running = False
        self._is_stopping = False
        self._close_after_stop = False
        self.observer = None
        self.executor = None
        self.excel_logger = None
        self.invoice_matcher = None
        self.tax_processor = None
        self.ledger_sync_manager = None
        self.operation_manifest = None
        self.stop_event = threading.Event()
        self.pending_set: set = set()
        self._log_queue: queue.Queue = queue.Queue()
        self._finance_batch_running = False

        self._build_ui()
        self._poll_log_queue()

    # ── UI 构建 ──────────────────────────────

    def _build_ui(self):
        # 顶部标题
        title_frame = tk.Frame(self, bg="#2E86AB", pady=12)
        title_frame.pack(fill="x")
        tk.Label(
            title_frame, text="📁 文件自动归类助手",
            font=("微软雅黑", 18, "bold"), fg="white", bg="#2E86AB"
        ).pack()
        tk.Label(
            title_frame,
            text="实时监视文件夹 · 按发票号归类 · 判断退税材料 · 自动改名",
            font=("微软雅黑", 10), fg="#D6EAF8", bg="#2E86AB"
        ).pack()

        # 主内容区
        content = tk.Frame(self, bg="#F0F4F8", padx=20, pady=10)
        content.pack(fill="both", expand=True)

        # 文件夹选择区
        folder_frame = tk.LabelFrame(
            content, text="  📂 文件夹设置  ",
            font=("微软雅黑", 11, "bold"), bg="#F0F4F8",
            fg="#2C3E50", padx=10, pady=8
        )
        folder_frame.pack(fill="x", pady=(0, 10))

        # 监控文件夹
        self._folder_row(folder_frame, "监控文件夹（待归类文件放这里）：", self.watch_dir,
                         self._choose_watch, row=0)
        # 输出文件夹
        self._folder_row(folder_frame, "输出主文件夹（归类后存放位置）：", self.output_dir,
                         self._choose_output, row=1)

        safety_frame = tk.LabelFrame(
            content, text="  安全选项  ",
            font=("微软雅黑", 10, "bold"), bg="#F0F4F8",
            fg="#2C3E50", padx=10, pady=6
        )
        safety_frame.pack(fill="x", pady=(0, 10))

        tk.Checkbutton(
            safety_frame,
            text="预览模式：只生成计划，不复制、不改名",
            variable=self.dry_run_var,
            font=("微软雅黑", 10),
            bg="#F0F4F8",
            fg="#2C3E50",
            activebackground="#F0F4F8",
        ).pack(side="left", padx=(0, 18))

        tk.Checkbutton(
            safety_frame,
            text="生成标准命名副本",
            variable=self.auto_rename_var,
            font=("微软雅黑", 10),
            bg="#F0F4F8",
            fg="#2C3E50",
            activebackground="#F0F4F8",
        ).pack(side="left")

        # 控制按钮区
        btn_frame = tk.Frame(content, bg="#F0F4F8")
        btn_frame.pack(fill="x", pady=(0, 10))

        self.btn_start = tk.Button(
            btn_frame, text="▶  开始监控", font=("微软雅黑", 12, "bold"),
            bg="#27AE60", fg="white", activebackground="#1E8449",
            relief="flat", padx=20, pady=8, cursor="hand2",
            command=self._start
        )
        self.btn_start.pack(side="left", padx=(0, 10))

        self.btn_stop = tk.Button(
            btn_frame, text="⏹  停止监控", font=("微软雅黑", 12, "bold"),
            bg="#E74C3C", fg="white", activebackground="#C0392B",
            relief="flat", padx=20, pady=8, cursor="hand2",
            state="disabled", command=self._stop
        )
        self.btn_stop.pack(side="left", padx=(0, 10))

        self.btn_finance_batch = tk.Button(
            btn_frame, text="📦  生成财务批次包", font=("微软雅黑", 11, "bold"),
            bg="#8E44AD", fg="white", activebackground="#6C3483",
            relief="flat", padx=14, pady=8, cursor="hand2",
            command=self._generate_finance_batch
        )
        self.btn_finance_batch.pack(side="left", padx=(0, 10))

        self.btn_undo = tk.Button(
            btn_frame, text="↩  撤销上次整理", font=("微软雅黑", 11, "bold"),
            bg="#D35400", fg="white", activebackground="#A04000",
            relief="flat", padx=14, pady=8, cursor="hand2",
            command=self._undo_last_operations
        )
        self.btn_undo.pack(side="left", padx=(0, 10))

        self.btn_clear = tk.Button(
            btn_frame, text="🗑  清空日志", font=("微软雅黑", 11),
            bg="#95A5A6", fg="white", activebackground="#7F8C8D",
            relief="flat", padx=14, pady=8, cursor="hand2",
            command=self._clear_log
        )
        self.btn_clear.pack(side="left")

        # 状态栏
        self.status_var = tk.StringVar(value="⚫ 未启动")
        status_bar = tk.Label(
            btn_frame, textvariable=self.status_var,
            font=("微软雅黑", 11), bg="#F0F4F8", fg="#7F8C8D"
        )
        status_bar.pack(side="right", padx=10)

        # 统计信息
        stats_frame = tk.Frame(content, bg="#F0F4F8")
        stats_frame.pack(fill="x", pady=(0, 6))
        self.stats_var = tk.StringVar(value="已处理文件：0 个")
        tk.Label(
            stats_frame, textvariable=self.stats_var,
            font=("微软雅黑", 10), bg="#F0F4F8", fg="#2980B9"
        ).pack(side="left")

        # 日志区
        log_frame = tk.LabelFrame(
            content, text="  📋 运行日志  ",
            font=("微软雅黑", 11, "bold"), bg="#F0F4F8",
            fg="#2C3E50", padx=6, pady=6
        )
        log_frame.pack(fill="both", expand=True)

        self.log_text = scrolledtext.ScrolledText(
            log_frame, font=("Consolas", 9), bg="#1E2A35", fg="#ECF0F1",
            insertbackground="white", state="disabled",
            wrap="word", relief="flat"
        )
        self.log_text.pack(fill="both", expand=True)

        # 底部提示
        tip = tk.Label(
            self, text="💡 默认预览模式不会改动文件；确认计划无误后取消勾选再执行。所有真实复制都会写入操作清单，可用【撤销上次整理】回退。",
            font=("微软雅黑", 9), bg="#D5E8D4", fg="#27AE60", pady=4
        )
        tip.pack(fill="x", side="bottom")

        self._processed_count = 0

    def _folder_row(self, parent, label_text, var, cmd, row):
        tk.Label(
            parent, text=label_text,
            font=("微软雅黑", 10), bg="#F0F4F8", fg="#2C3E50"
        ).grid(row=row, column=0, sticky="w", pady=4)

        entry = tk.Entry(
            parent, textvariable=var,
            font=("微软雅黑", 10), width=42,
            relief="solid", bd=1
        )
        entry.grid(row=row, column=1, padx=8, pady=4, sticky="ew")

        tk.Button(
            parent, text="浏览…", font=("微软雅黑", 10),
            bg="#3498DB", fg="white", activebackground="#2980B9",
            relief="flat", padx=8, cursor="hand2", command=cmd
        ).grid(row=row, column=2, padx=(0, 4), pady=4)

        parent.columnconfigure(1, weight=1)

    # ── 事件处理 ─────────────────────────────

    def _choose_watch(self):
        d = filedialog.askdirectory(title="选择监控文件夹")
        if d:
            self.watch_dir.set(d)

    def _choose_output(self):
        d = filedialog.askdirectory(title="选择输出主文件夹")
        if d:
            self.output_dir.set(d)

    def _undo_last_operations(self):
        if self.is_running or self._is_stopping:
            messagebox.showwarning("提示", "请先停止监控，再执行撤销。")
            return
        output = self.output_dir.get().strip()
        if not output or not os.path.isdir(output):
            messagebox.showwarning("提示", "请先选择有效的输出主文件夹！")
            return
        manifest_path = latest_operation_manifest(normalize_path(output))
        if not manifest_path:
            messagebox.showinfo("提示", "没有找到可撤销的操作清单。")
            return
        if not messagebox.askyesno(
            "确认撤销",
            "将根据最近一次操作清单删除本程序复制生成的文件。\n"
            "如果目标文件已被人工修改，会自动跳过。\n\n"
            f"操作清单：\n{manifest_path}\n\n"
            "确定继续吗？",
        ):
            return

        self.btn_undo.config(state="disabled")
        self._enqueue_log(f"↩ 开始撤销: {manifest_path}")

        def worker():
            try:
                summary = undo_operation_manifest(manifest_path)
                self._log_queue.put(("undo_done", summary))
            except Exception as exc:
                self._log_queue.put(("undo_error", str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _count_candidate_files(self, watch: str, output: str, invoice_matcher: InvoiceMatcher) -> int:
        count = 0
        for root, _, files in os.walk(watch):
            for filename in files:
                filepath = normalize_path(os.path.join(root, filename))
                if should_ignore_file(filepath) or is_path_inside(filepath, output):
                    continue
                if invoice_matcher.match_path(filepath, source_root=watch):
                    count += 1
        return count

    def _confirm_real_run(self, candidate_count: int) -> bool:
        if self.dry_run_var.get():
            return True
        auto_rename_text = (
            "\n已启用【生成标准命名副本】，不会移动原文件夹，但会复制生成新文件夹。"
            if self.auto_rename_var.get()
            else ""
        )
        if candidate_count > SAFE_BATCH_CONFIRM_LIMIT:
            value = simpledialog.askstring(
                "二次确认",
                f"当前将真实复制 {candidate_count} 个已有文件，超过安全阈值 {SAFE_BATCH_CONFIRM_LIMIT}。"
                f"{auto_rename_text}\n\n如确认执行，请输入：确认执行",
                parent=self,
            )
            return (value or "").strip() == "确认执行"
        return messagebox.askyesno(
            "确认执行",
            f"当前不是预览模式，将真实复制 {candidate_count} 个已有文件。"
            f"{auto_rename_text}\n\n确认继续吗？",
        )

    def _generate_finance_batch(self):
        if self._finance_batch_running:
            return
        if openpyxl is None:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\npip install openpyxl")
            return

        output = self.output_dir.get().strip()
        if not output or not os.path.isdir(output):
            messagebox.showwarning("提示", "请先选择有效的输出主文件夹！")
            return
        output = normalize_path(output)

        if self.is_running:
            messagebox.showinfo(
                "提示",
                "当前正在监控文件夹。财务批次处理会复制文件夹，不会停止监控；"
                "如果 Excel 或材料文件正在被占用，可能需要关闭后重试。",
            )

        batch_path = filedialog.askopenfilename(
            title="选择财务批次发票号 Excel",
            filetypes=[("Excel 工作簿", "*.xlsx")],
        )
        if not batch_path:
            return

        self._finance_batch_running = True
        self.btn_finance_batch.config(state="disabled")
        self._enqueue_log(f"📦 已选择财务批次表: {batch_path}")

        def worker():
            try:
                settings = ensure_app_settings(APP_SETTINGS_PATH)
                invoice_workbook = find_invoice_workbook(settings)
                if not invoice_workbook:
                    raise RuntimeError(
                        "未找到公司系统出运统计发票号清单，请把主台账放到："
                        f"{DEFAULT_COMPANY_LEDGER}"
                    )
                if self.tax_processor:
                    tax_processor = self.tax_processor
                else:
                    invoice_records = load_invoice_records(str(invoice_workbook))
                    tax_processor = TaxRefundProcessor(output, ledger_records=invoice_records)

                processor = FinanceBatchProcessor(
                    output_dir=output,
                    company_ledger_path=str(invoice_workbook),
                    tax_processor=tax_processor,
                    settings=settings,
                    log_callback=self._enqueue_log,
                )
                summary = processor.process_batch(batch_path)
                self._log_queue.put(("finance_batch_done", summary))
            except Exception as exc:
                self._log_queue.put(("finance_batch_error", str(exc)))

        threading.Thread(target=worker, daemon=True).start()

    def _start(self):
        if Observer is None:
            messagebox.showerror("缺少依赖", "请先安装 watchdog：\npip install watchdog openpyxl")
            return
        if openpyxl is None:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\npip install openpyxl")
            return

        watch = self.watch_dir.get().strip()
        output = self.output_dir.get().strip()

        if not watch or not os.path.isdir(watch):
            messagebox.showwarning("提示", "请先选择有效的监控文件夹！")
            return
        if not output:
            messagebox.showwarning("提示", "请先选择输出主文件夹！")
            return

        watch = normalize_path(watch)
        output = normalize_path(output)
        os.makedirs(output, exist_ok=True)

        if paths_equal(watch, output):
            messagebox.showwarning("提示", "监控文件夹和输出文件夹不能相同！")
            return
        if is_path_inside(output, watch):
            messagebox.showwarning("提示", "输出文件夹不能放在监控文件夹里面！")
            return

        settings = ensure_app_settings(APP_SETTINGS_PATH)
        invoice_workbook = find_invoice_workbook(settings)
        if not invoice_workbook:
            messagebox.showwarning(
                "提示",
                "未找到公司系统出运统计发票号清单。\n"
                f"请把主台账放到：\n{DEFAULT_COMPANY_LEDGER}",
            )
            return
        try:
            invoice_records = load_invoice_records(str(invoice_workbook))
            invoice_numbers = sorted(invoice_records.keys(), key=len, reverse=True)
        except Exception as exc:
            messagebox.showerror("发票号清单读取失败", str(exc))
            return
        if not invoice_numbers:
            messagebox.showwarning("提示", f"公司系统主台账没有可用发票号：\n{invoice_workbook}")
            return

        preview_matcher = InvoiceMatcher(invoice_numbers)
        existing_count = self._count_candidate_files(watch, output, preview_matcher)
        if not self._confirm_real_run(existing_count):
            return

        dry_run = self.dry_run_var.get()
        auto_rename = self.auto_rename_var.get()
        run_id = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.operation_manifest = OperationManifest(output, dry_run=dry_run, run_id=run_id)

        self.is_running = True
        self._is_stopping = False
        self._close_after_stop = False
        self._processed_count = 0
        self.stats_var.set("已处理文件：0 个")
        self.pending_set.clear()
        self.stop_event = threading.Event()
        self.watch_dir.set(watch)
        self.output_dir.set(output)

        # OCR and PDF parsing are CPU-heavy, especially in Windows VMs.
        # Keep concurrency bounded so the UI remains usable during batch drops.
        self.executor = ThreadPoolExecutor(max_workers=MAX_FILE_PROCESS_WORKERS)

        # 创建 Excel 日志管理器
        self.excel_logger = ExcelLogger(output)
        self.invoice_matcher = preview_matcher
        try:
            self.tax_processor = TaxRefundProcessor(
                output,
                ledger_records=invoice_records,
                dry_run=dry_run,
                operation_manifest=self.operation_manifest,
            )
        except Exception as exc:
            messagebox.showerror("退税材料配置读取失败", str(exc))
            self.executor.shutdown(wait=False, cancel_futures=True)
            self.executor = None
            self.is_running = False
            return
        self.ledger_sync_manager = CompanyLedgerSyncManager(
            str(invoice_workbook),
            batch_name=settings.get("finance_batch", {}).get("last_batch_name", ""),
            log_callback=self._enqueue_log,
            debounce_seconds=5.0,
        )

        # 创建文件监控
        handler = FileHandler(
            output_dir=output,
            excel_logger=self.excel_logger,
            invoice_matcher=self.invoice_matcher,
            executor=self.executor,
            log_callback=self._enqueue_log,
            pending_set=self.pending_set,
            tax_processor=self.tax_processor,
            declaration_prompt_callback=self._request_declaration_no,
            company_ledger_path=str(invoice_workbook),
            finance_batch_name=settings.get("finance_batch", {}).get("last_batch_name", ""),
            ledger_sync_manager=self.ledger_sync_manager,
            watch_dir=watch,
            stop_event=self.stop_event,
            dry_run=dry_run,
            operation_manifest=self.operation_manifest,
            auto_rename=auto_rename,
        )

        self.observer = Observer()
        self.observer.schedule(handler, watch, recursive=True)
        self.observer.start()
        submitted_count = self._scan_existing_files(watch, handler)

        self._enqueue_log(f"🚀 开始监控: {watch}")
        self._enqueue_log(f"📂 输出目录: {output}")
        self._enqueue_log(f"🧾 操作清单: {self.operation_manifest.path}")
        if dry_run:
            self._enqueue_log("🧾 当前为预览模式：不会复制、改名或移动文件")
        else:
            self._enqueue_log("⚠ 当前为真实执行模式：所有复制都会写入操作清单，可撤销")
        self._enqueue_log(f"📒 发票号清单: {invoice_workbook}（{len(invoice_numbers)} 个发票号）")
        self._enqueue_log(f"📋 退税材料规则: {DEFAULT_MATERIAL_CONFIG}")
        self._enqueue_log(f"🔁 已启用递归监听；启动时已提交 {submitted_count} 个已有文件")
        self._enqueue_log("─" * 60)

        self.executor.submit(
            self.tax_processor.process_all_invoice_folders,
            invoice_numbers,
            self._request_declaration_no,
            self._enqueue_log,
            auto_rename,
        )

        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self.status_var.set("🟢 监控中…")

    def _scan_existing_files(self, watch: str, handler: FileHandler) -> int:
        submitted = 0
        for root, _, files in os.walk(watch):
            for filename in files:
                filepath = os.path.join(root, filename)
                if handler._schedule(filepath):
                    submitted += 1
        return submitted

    def _stop(self):
        if self._is_stopping:
            return
        if not self.observer and not self.executor:
            return

        self._is_stopping = True
        if self.stop_event:
            self.stop_event.set()
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="disabled")
        self.status_var.set("🟡 正在停止…")
        self._enqueue_log("─" * 60)
        self._enqueue_log("⏳ 正在停止监控，取消未开始任务并中止待处理文件…")

        observer = self.observer
        executor = self.executor
        excel_logger = self.excel_logger
        ledger_sync_manager = self.ledger_sync_manager
        self.observer = None
        self.executor = None
        self.excel_logger = None

        def shutdown_worker():
            errors = []
            try:
                if observer:
                    observer.stop()
                    observer.join(timeout=5)
            except Exception as exc:
                errors.append(f"停止文件监听失败: {exc}")

            try:
                if executor:
                    executor.shutdown(wait=False, cancel_futures=True)
            except Exception as exc:
                errors.append(f"取消处理线程失败: {exc}")

            try:
                if excel_logger:
                    excel_logger.save()
            except Exception as exc:
                errors.append(f"保存工作日志失败: {exc}")

            try:
                if ledger_sync_manager:
                    ledger_sync_manager.close(flush=True)
            except Exception as exc:
                errors.append(f"同步公司主台账失败: {exc}")

            self._log_queue.put(("stop_done", errors))

        threading.Thread(target=shutdown_worker, daemon=True).start()

    def _finish_stop(self, errors: list[str]):
        self.is_running = False
        self._is_stopping = False
        self.pending_set.clear()
        self.invoice_matcher = None
        self.tax_processor = None
        self.ledger_sync_manager = None

        for err in errors:
            self._write_log_line(f"❌ {err}")
        self._write_log_line(f"⏹ 已停止监控。本次共处理 {self._processed_count} 个文件。")
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.btn_undo.config(state="normal")
        self.status_var.set("⚫ 已停止")

        if self._close_after_stop:
            self.destroy()

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.config(state="disabled")

    def _enqueue_log(self, msg: str, processed: bool = False):
        self._log_queue.put(("log", msg, processed))

    def _request_declaration_no(self, invoice_no: str) -> str:
        event = threading.Event()
        payload = {
            "invoice_no": invoice_no,
            "value": "",
            "event": event,
        }
        self._log_queue.put(("prompt_declaration", payload))
        event.wait(DEFAULT_DECLARATION_TIMEOUT_SECONDS)
        return payload.get("value", "").strip()

    def _finish_finance_batch(self, summary: dict):
        self._finance_batch_running = False
        self.btn_finance_batch.config(state="normal")
        counts = summary.get("counts", {})
        message = (
            f"批次总数：{summary.get('total', 0)}\n"
            f"齐了：{counts.get('齐了', 0)}\n"
            f"缺材料：{counts.get('缺材料', 0)}\n"
            f"待复核：{counts.get('待复核', 0)}\n"
            f"未找到文件夹：{counts.get('未找到文件夹', 0)}\n"
            f"不在主台账：{counts.get('不在主台账', 0)}\n"
            f"批次输出目录：\n{summary.get('batch_output_dir', '')}"
        )
        messagebox.showinfo("财务批次处理完成", message)

    def _finish_finance_batch_error(self, error: str):
        self._finance_batch_running = False
        self.btn_finance_batch.config(state="normal")
        self._write_log_line(f"❌ 财务批次处理失败: {error}")
        messagebox.showerror("财务批次处理失败", error)

    def _write_log_line(self, msg: str):
        self.log_text.config(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.config(state="disabled")

    def _poll_log_queue(self):
        """每100ms从队列中取出日志并写入文本框"""
        try:
            while True:
                item = self._log_queue.get_nowait()
                if isinstance(item, tuple) and item and item[0] == "stop_done":
                    self._finish_stop(item[1])
                    continue
                if isinstance(item, tuple) and item and item[0] == "finance_batch_done":
                    self._finish_finance_batch(item[1])
                    continue
                if isinstance(item, tuple) and item and item[0] == "finance_batch_error":
                    self._finish_finance_batch_error(item[1])
                    continue
                if isinstance(item, tuple) and item and item[0] == "undo_done":
                    summary = item[1]
                    self.btn_undo.config(state="normal")
                    self._write_log_line(
                        f"↩ 撤销完成：已回退 {summary.get('undone', 0)} 项，"
                        f"跳过 {summary.get('skipped', 0)} 项，错误 {len(summary.get('errors', []))} 项"
                    )
                    if summary.get("errors"):
                        self._write_log_line("❌ 撤销错误：" + "；".join(summary["errors"]))
                    messagebox.showinfo(
                        "撤销完成",
                        f"已回退：{summary.get('undone', 0)} 项\n"
                        f"已跳过：{summary.get('skipped', 0)} 项\n"
                        f"错误：{len(summary.get('errors', []))} 项",
                    )
                    continue
                if isinstance(item, tuple) and item and item[0] == "undo_error":
                    self.btn_undo.config(state="normal")
                    self._write_log_line(f"❌ 撤销失败: {item[1]}")
                    messagebox.showerror("撤销失败", item[1])
                    continue
                if isinstance(item, tuple) and item and item[0] == "prompt_declaration":
                    payload = item[1]
                    invoice_no = payload.get("invoice_no", "")
                    value = simpledialog.askstring(
                        "输入报关号",
                        f"未能自动识别 {invoice_no} 的报关号。\n请输入报关号：",
                        parent=self,
                    )
                    payload["value"] = (value or "").strip()
                    payload["event"].set()
                    continue
                if isinstance(item, tuple) and item and item[0] == "log":
                    _, msg, processed = item
                else:
                    msg = str(item)
                    processed = False

                if processed:
                    self._processed_count += 1
                    self.stats_var.set(f"已处理文件：{self._processed_count} 个")
                self._write_log_line(msg)
        except queue.Empty:
            pass
        try:
            if self.winfo_exists():
                self.after(100, self._poll_log_queue)
        except tk.TclError:
            pass

    def on_closing(self):
        if self._is_stopping:
            self._close_after_stop = True
            self.status_var.set("🟡 正在停止，完成后关闭…")
        elif self.is_running:
            if messagebox.askyesno("确认退出", "监控正在运行，确定要退出吗？"):
                self._close_after_stop = True
                self._stop()
        else:
            self.destroy()


# ─────────────────────────────────────────────
# Shell E2E test-mode
# ─────────────────────────────────────────────

E2E_REPORT_SCHEMA = "contract-router-e2e/v1"


def e2e_fixture_texts() -> dict[str, str]:
    return {
        "random_a.pdf": (
            "PROFORMA INVOICE\n"
            "Invoice No. 25117260\n"
            "JIAXING LAYO\n"
            "Transport details\n"
            "Terms of payment\n"
            "Description of goods\n"
            "POLAND\n"
            "USD 100.00"
        ),
        "scan001.pdf": (
            "报关委托书\n"
            "代理报关委托\n"
            "报关单编号 123456789012345\n"
            "一般贸易\n"
            "177FCWCWS64050\n"
            "7"
        ),
        "20260212.pdf": (
            "电子发票\n"
            "发票号码 12345678\n"
            "国内货物运输代理服务\n"
            "报关\n"
            "嘉兴新锴国际货运代理\n"
            "177FCWCWS640507"
        ),
        "docx_export.pdf": (
            "OCEAN BILL OF LADING\n"
            "SHIPPER\n"
            "CONSIGNEE\n"
            "NOTIFY PARTY\n"
            "FREIGHT COLLECT\n"
            "SHIPPED ON BOARD\n"
            "SHIPPER'S LOAD COUNT\n"
            "B/L No. ESHGDN2600603\n"
            "Container MSBU5039559 Seal FX45652534\n"
            "12 CARTONS\n"
            "1234.56KGS\n"
            "MAY 20, 2026"
        ),
        "wechat_image.png": (
            "CONTAINER LOAD PLAN\n"
            "Container No. MSBU5039559\n"
            "Seal No. FX45652534\n"
            "Port of Loading SHANGHAI\n"
            "Port of Discharge GDANSK\n"
            "Place of Delivery GDANSK\n"
            "Packing Date 2026-05-19\n"
            "Total Packages 12\n"
            "Total Cargo Wt 1234.56KGS\n"
            "Total Meas 10\n"
            "40HC\n"
            "Bill of Lading No. 177FCWCWS64050\n"
            "7"
        ),
        "invoice_unknown.pdf": (
            "电子发票\n"
            "发票号码 87654321\n"
            "国际货运代理费\n"
            "港口码头费\n"
            "汇利达欧海国际货运代理\n"
            "B/L No. ESHGDN2600603"
        ),
        "duplicate_export_sales.pdf": (
            "PROFORMA INVOICE\n"
            "Invoice No. LY25117260\n"
            "JIAXING LAYO\n"
            "Transport details\n"
            "Terms of payment\n"
            "Description of goods\n"
            "POLAND\n"
            "USD 100.00"
        ),
    }


def _e2e_save_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    if openpyxl is None:
        raise RuntimeError("缺少 openpyxl，无法生成 E2E 台账 fixture")
    wb = openpyxl.Workbook()
    first = True
    try:
        for title, rows in sheets.items():
            ws = wb.active if first else wb.create_sheet(title)
            ws.title = title
            first = False
            for row in rows:
                ws.append(row)
        path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(path)
    finally:
        wb.close()


def _e2e_write_dummy_files(folder: Path, filenames: list[str]) -> list[Path]:
    folder.mkdir(parents=True, exist_ok=True)
    paths = []
    for filename in filenames:
        path = folder / filename
        path.write_bytes(b"contract-router e2e fixture\n")
        paths.append(path)
    return paths


def _e2e_purchase_contract_filenames(invoice_digits_value: str, count: int) -> list[str]:
    factories = ["优尚", "莱尔特", "普沃特", "优创", "嘉结", "海泉", "凯艺美", "华鸣", "协作工厂A", "协作工厂B"]
    return [
        f"采购合同-{factories[index % len(factories)]}-{invoice_digits_value}-{index + 1:02d}.jpg"
        for index in range(count)
    ]


def _e2e_split_marker_text(value: object) -> list[str]:
    if not value:
        return []
    return [item.strip() for item in re.split(r"[、；;,]+", str(value)) if item.strip()]


def _e2e_markers_from_record(record: dict) -> dict[str, list[str]]:
    return {
        "conflict": _e2e_split_marker_text(record.get("闭环冲突", "")),
        "duplicate": _e2e_split_marker_text(record.get("重复材料", "")),
        "unmatched": _e2e_split_marker_text(record.get("未识别文件", "")),
    }


def _e2e_assert(assertions: list[dict], name: str, passed: bool, details: object = "") -> None:
    assertions.append({
        "name": name,
        "status": "passed" if passed else "failed",
        "details": details,
    })


def _e2e_case_status(assertions: list[dict]) -> str:
    return "passed" if all(item["status"] == "passed" for item in assertions) else "failed"


def _e2e_record_summary(record: dict) -> dict:
    keys = [
        "发票号", "最终发票号显示值", "报关号", "状态", "已识别必需材料", "缺少材料",
        "未识别文件", "重复材料", "需人工确认文件", "闭环状态", "闭环分数",
        "闭环证据", "闭环冲突", "闭环未核验",
    ]
    return {key: record.get(key, "") for key in keys}


def _run_e2e_random_filename_case(
    fixture_root: Path,
    output_root: Path,
    config_path: Path,
    ledger_records: dict,
    logs: list[str],
) -> dict:
    assertions = []
    source_root = fixture_root / "random_filenames"
    incoming_dir = source_root / "25117260 raw"
    filenames = [
        "random_a.pdf",
        "scan001.pdf",
        "20260212.pdf",
        "docx_export.pdf",
        "wechat_image.png",
        "invoice_unknown.pdf",
    ]
    source_files = _e2e_write_dummy_files(incoming_dir, filenames)
    source_files.extend(_e2e_write_dummy_files(source_root, _e2e_purchase_contract_filenames("25117260", 6)))

    processor = TaxRefundProcessor(
        str(output_root / "random_filenames"),
        config_path=config_path,
        ledger_records=ledger_records,
    )
    matcher = InvoiceMatcher(list(ledger_records.keys()))
    moves = []
    for source in source_files:
        success, subfolder, message, dest_path, invoice_no, status = move_file_to_output(
            str(source),
            processor.output_dir,
            matcher,
            invoice_dir_resolver=processor.resolve_target_folder,
            source_root=str(source_root),
        )
        moves.append({
            "source": str(source),
            "success": success,
            "subfolder": subfolder,
            "message": message,
            "dest_path": dest_path,
            "invoice_no": invoice_no,
            "status": status,
        })

    invoice_folder = Path(processor.resolve_target_folder("LY25117260"))
    record = processor.evaluate_invoice(
        "LY25117260",
        str(invoice_folder),
        allow_prompt=False,
        allow_rename=False,
        save_report=False,
        persist_declaration=False,
    )
    evidence = record.get("闭环证据", "")
    markers = _e2e_markers_from_record(record)

    _e2e_assert(assertions, "all_random_named_files_match_same_invoice", all(
        item["success"] and item["invoice_no"] == "LY25117260" for item in moves
    ), moves)
    _e2e_assert(assertions, "ly_prefix_and_digits_are_equivalent", invoice_equivalent(
        "LY25117260",
        record.get("最终发票号显示值", ""),
    ), record.get("最终发票号显示值", ""))
    _e2e_assert(assertions, "closed_loop_strong_pass", record.get("闭环状态") == "强闭环通过", _e2e_record_summary(record))
    _e2e_assert(assertions, "customs_declaration_detected", record.get("报关号") == "123456789012345", record.get("报关号", ""))
    _e2e_assert(assertions, "amount_country_date_evidence_present", all(
        marker in evidence for marker in ["台账金额", "台账目的国", "台账出运日期"]
    ), evidence)
    _e2e_assert(assertions, "no_conflict_duplicate_or_unmatched_in_happy_path", not any(markers.values()), markers)

    logs.append(f"random_filenames status={record.get('状态')} closed_loop={record.get('闭环状态')}")
    return {
        "name": "random_filenames_closed_loop",
        "status": _e2e_case_status(assertions),
        "assertions": assertions,
        "markers": markers,
        "record": _e2e_record_summary(record),
        "moves": moves,
    }


def _run_e2e_amount_conflict_case(
    output_root: Path,
    config_path: Path,
    text_map: dict[str, str],
    logs: list[str],
) -> dict:
    assertions = []
    folder = output_root / "amount_conflict" / "LY25117260"
    _e2e_write_dummy_files(folder, [
        "random_a.pdf",
        "scan001.pdf",
        "20260212.pdf",
        "docx_export.pdf",
        "wechat_image.png",
        "invoice_unknown.pdf",
    ])
    ledger_records = {
        "LY25117260": {
            "发票号": "LY25117260",
            "金额": 999,
            "目的国": "波兰",
            "出运日期": "2026-05-20",
        }
    }
    processor = TaxRefundProcessor(
        str(output_root / "amount_conflict"),
        config_path=config_path,
        ledger_records=ledger_records,
    )
    record = processor.evaluate_invoice(
        "LY25117260",
        str(folder),
        allow_prompt=False,
        allow_rename=False,
        save_report=False,
        persist_declaration=False,
    )
    markers = _e2e_markers_from_record(record)

    _e2e_assert(assertions, "amount_conflict_marked", any(
        "金额" in item for item in markers["conflict"]
    ), markers)
    _e2e_assert(assertions, "amount_conflict_not_accepted", record.get("状态") not in {"可改名", "已改名"}, _e2e_record_summary(record))
    _e2e_assert(assertions, "conflict_fixture_uses_same_invoice_content", invoice_equivalent(
        "LY25117260",
        text_map["random_a.pdf"].split("Invoice No.", 1)[1].splitlines()[0].strip(),
    ), text_map["random_a.pdf"])

    logs.append(f"amount_conflict status={record.get('状态')} conflicts={record.get('闭环冲突')}")
    return {
        "name": "amount_conflict",
        "status": _e2e_case_status(assertions),
        "assertions": assertions,
        "markers": markers,
        "record": _e2e_record_summary(record),
    }


def _run_e2e_duplicate_case(
    output_root: Path,
    config_path: Path,
    ledger_records: dict,
    logs: list[str],
) -> dict:
    assertions = []
    folder = output_root / "duplicate_invoice_material" / "LY25117260"
    _e2e_write_dummy_files(folder, [
        "random_a.pdf",
        "duplicate_export_sales.pdf",
        "scan001.pdf",
        "20260212.pdf",
        "docx_export.pdf",
        "wechat_image.png",
        "invoice_unknown.pdf",
    ])
    processor = TaxRefundProcessor(
        str(output_root / "duplicate_invoice_material"),
        config_path=config_path,
        ledger_records=ledger_records,
    )
    record = processor.evaluate_invoice(
        "LY25117260",
        str(folder),
        allow_prompt=False,
        allow_rename=False,
        save_report=False,
        persist_declaration=False,
    )
    markers = _e2e_markers_from_record(record)
    warnings = []
    if markers["duplicate"] and record.get("状态") in {"可改名", "已改名"}:
        warnings.append("duplicate marker exists, but application record status is still accepted")

    _e2e_assert(assertions, "duplicate_invoice_material_marked", bool(markers["duplicate"]), markers)
    _e2e_assert(assertions, "duplicate_marker_names_invoice_file", any(
        name in {"random_a.pdf", "duplicate_export_sales.pdf"} for name in markers["duplicate"]
    ), markers)

    logs.append(f"duplicate_invoice_material status={record.get('状态')} duplicate={record.get('重复材料')}")
    return {
        "name": "duplicate_invoice_material",
        "status": _e2e_case_status(assertions),
        "assertions": assertions,
        "warnings": warnings,
        "markers": markers,
        "record": _e2e_record_summary(record),
    }


def _run_e2e_unmatched_case(
    fixture_root: Path,
    output_root: Path,
    ledger_records: dict,
    logs: list[str],
) -> dict:
    assertions = []
    source_root = fixture_root / "unmatched"
    source_file = _e2e_write_dummy_files(source_root, ["totally_unrelated_file.pdf"])[0]
    matcher = InvoiceMatcher(list(ledger_records.keys()))
    success, subfolder, message, dest_path, invoice_no, status = move_file_to_output(
        str(source_file),
        str(output_root / "unmatched"),
        matcher,
        source_root=str(source_root),
    )
    event = {
        "source": str(source_file),
        "success": success,
        "subfolder": subfolder,
        "message": message,
        "dest_path": dest_path,
        "invoice_no": invoice_no,
        "status": status,
    }
    markers = {
        "conflict": [],
        "duplicate": [],
        "unmatched": [Path(dest_path).name] if success and status == "未匹配" else [],
    }

    _e2e_assert(assertions, "unrelated_file_marked_unmatched", bool(markers["unmatched"]), event)
    _e2e_assert(assertions, "unrelated_file_goes_to_review_folder", subfolder == REVIEW_SUBFOLDER, event)
    _e2e_assert(assertions, "unrelated_file_has_no_invoice_match", invoice_no == "", event)

    logs.append(f"unmatched status={status} dest={dest_path}")
    return {
        "name": "unrelated_file_unmatched",
        "status": _e2e_case_status(assertions),
        "assertions": assertions,
        "markers": markers,
        "moves": [event],
    }


def _write_e2e_report(report_path: Path, report: dict) -> None:
    report_path.parent.mkdir(parents=True, exist_ok=True)
    with open(report_path, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)


def run_contract_router_e2e(
    input_dir: str | Path,
    output_dir: str | Path,
    report_path: str | Path,
    log_path: str | Path | None = None,
) -> int:
    started_at = datetime.now().isoformat(timespec="seconds")
    input_dir = Path(input_dir)
    output_dir = Path(output_dir)
    report_path = Path(report_path)
    log_path = Path(log_path) if log_path else output_dir / "logs" / "e2e.log"
    run_id = f"run-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{os.getpid()}"
    fixture_root = input_dir / run_id
    run_output = output_dir / run_id
    logs = [
        f"started_at={started_at}",
        f"run_id={run_id}",
        f"fixture_root={fixture_root}",
        f"run_output={run_output}",
    ]

    original_pdf_text = globals()["pdf_text"]
    original_ocr_text = globals()["ocr_text"]
    try:
        fixture_root.mkdir(parents=True, exist_ok=True)
        run_output.mkdir(parents=True, exist_ok=True)
        text_map = e2e_fixture_texts()

        def fixture_text(filepath: str, *args, **kwargs) -> str:
            del args, kwargs
            return text_map.get(Path(filepath).name, "")

        globals()["pdf_text"] = fixture_text
        globals()["ocr_text"] = fixture_text

        ledger_path = fixture_root / "ledger.xlsx"
        _e2e_save_workbook(
            ledger_path,
            {
                "台账": [
                    ["发票号", "出运日期", "目的国(地区)", "金额"],
                    ["LY25117260", "2026-05-20", "波兰", 100],
                    ["LY25117261", "2026-05-21", "德国", 200],
                ],
            },
        )
        ledger_records = load_invoice_records(str(ledger_path))
        config_path = fixture_root / "materials_config.json"

        cases = [
            _run_e2e_random_filename_case(
                fixture_root,
                run_output,
                config_path,
                ledger_records,
                logs,
            ),
            _run_e2e_amount_conflict_case(
                run_output,
                config_path,
                text_map,
                logs,
            ),
            _run_e2e_duplicate_case(
                run_output,
                config_path,
                ledger_records,
                logs,
            ),
            _run_e2e_unmatched_case(
                fixture_root,
                run_output,
                ledger_records,
                logs,
            ),
        ]
        marker_counts = {
            key: sum(len(case.get("markers", {}).get(key, [])) for case in cases)
            for key in ["conflict", "duplicate", "unmatched"]
        }
        warnings = [
            {"case": case["name"], "message": warning}
            for case in cases
            for warning in case.get("warnings", [])
        ]
        report_status = "passed" if all(case["status"] == "passed" for case in cases) else "failed"
        ended_at = datetime.now().isoformat(timespec="seconds")
        report = {
            "schema": E2E_REPORT_SCHEMA,
            "status": report_status,
            "started_at": started_at,
            "ended_at": ended_at,
            "run_id": run_id,
            "input": str(input_dir),
            "output": str(output_dir),
            "fixture_root": str(fixture_root),
            "run_output": str(run_output),
            "summary": {
                "total": len(cases),
                "passed": sum(1 for case in cases if case["status"] == "passed"),
                "failed": sum(1 for case in cases if case["status"] != "passed"),
                "markers": marker_counts,
                "warnings": len(warnings),
            },
            "warnings": warnings,
            "cases": cases,
        }
        _write_e2e_report(report_path, report)
        logs.append(f"ended_at={ended_at}")
        logs.append(f"status={report_status}")
        logs.append(f"report={report_path}")
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")
        print(f"[contract-router-e2e] report={report_path}")
        print(f"[contract-router-e2e] status={report_status}")
        return 0 if report_status == "passed" else 1
    except Exception as exc:
        ended_at = datetime.now().isoformat(timespec="seconds")
        report = {
            "schema": E2E_REPORT_SCHEMA,
            "status": "failed",
            "started_at": started_at,
            "ended_at": ended_at,
            "run_id": run_id,
            "input": str(input_dir),
            "output": str(output_dir),
            "fixture_root": str(fixture_root),
            "run_output": str(run_output),
            "summary": {
                "total": 0,
                "passed": 0,
                "failed": 1,
                "markers": {"conflict": 0, "duplicate": 0, "unmatched": 0},
                "warnings": 0,
            },
            "warnings": [],
            "error": {
                "type": type(exc).__name__,
                "message": str(exc),
            },
            "cases": [],
        }
        _write_e2e_report(report_path, report)
        logs.append(f"error={type(exc).__name__}: {exc}")
        log_path.parent.mkdir(parents=True, exist_ok=True)
        log_path.write_text("\n".join(logs) + "\n", encoding="utf-8")
        print(f"[contract-router-e2e] failed: {exc}", file=sys.stderr)
        print(f"[contract-router-e2e] report={report_path}")
        return 1
    finally:
        globals()["pdf_text"] = original_pdf_text
        globals()["ocr_text"] = original_ocr_text


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Contract Router")
    subparsers = parser.add_subparsers(dest="command", required=True)
    e2e_parser = subparsers.add_parser("e2e", help="Run shell-based E2E test-mode and write report.json.")
    e2e_parser.add_argument("--input", required=True, help="Fixture input/work directory.")
    e2e_parser.add_argument("--output", required=True, help="E2E output directory.")
    e2e_parser.add_argument("--report", required=True, help="Machine-readable report.json path.")
    e2e_parser.add_argument("--log", help="Optional E2E log path.")
    undo_parser = subparsers.add_parser("undo", help="Undo copied files from an operation manifest.")
    undo_parser.add_argument("--output", required=True, help="Output directory containing _操作日志.")
    undo_parser.add_argument("--manifest", help="Specific operation_manifest_*.json to undo. Defaults to latest.")
    undo_parser.add_argument("--dry-run", action="store_true", help="Preview undo without deleting files.")

    args = parser.parse_args(argv)
    if args.command == "e2e":
        return run_contract_router_e2e(args.input, args.output, args.report, args.log)
    if args.command == "undo":
        manifest_path = args.manifest or latest_operation_manifest(args.output)
        if not manifest_path:
            print(f"没有找到操作清单：{operation_manifest_dir(args.output)}", file=sys.stderr)
            return 1
        summary = undo_operation_manifest(manifest_path, dry_run=args.dry_run)
        print(f"manifest={summary['manifest_path']}")
        print(f"undone={summary['undone']} skipped={summary['skipped']} errors={len(summary['errors'])}")
        for error in summary["errors"]:
            print(f"error={error}", file=sys.stderr)
        return 0 if not summary["errors"] else 1
    parser.error("unknown command")
    return 2


# ─────────────────────────────────────────────
# 依赖检查与安装提示
# ─────────────────────────────────────────────

def check_and_install_deps():
    checks = [
        ("watchdog", "watchdog"),
        ("openpyxl", "openpyxl"),
        ("fitz", "PyMuPDF"),
        ("rapidocr_onnxruntime", "rapidocr-onnxruntime"),
    ]
    missing = []
    for module_name, package_name in checks:
        try:
            __import__(module_name)
        except ImportError:
            missing.append(package_name)

    if missing:
        import subprocess
        print(f"检测到缺少依赖: {', '.join(missing)}")
        print("正在自动安装……")
        for pkg in missing:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        print("依赖安装完成，请重新运行程序。")
        sys.exit(0)


# ─────────────────────────────────────────────
# 入口
# ─────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] in {"e2e", "undo", "-h", "--help"}:
        sys.exit(main(sys.argv[1:]))

    check_and_install_deps()

    # 重新导入（安装后）
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    app = App()
    app.protocol("WM_DELETE_WINDOW", app.on_closing)
    app.mainloop()
